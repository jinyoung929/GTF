import http.cookiejar
import json
import os
import socket
import subprocess
import sys
import time
import unittest
import urllib.error
import urllib.parse
import urllib.request

from openpyxl import load_workbook
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
PORT = 4191
BASE_URL = f"http://127.0.0.1:{PORT}"


def read_env_value(name: str) -> str:
    for env_path in (ROOT / ".env", ROOT / ".env.local"):
        if not env_path.exists():
            continue
        for line in env_path.read_text(encoding="utf-8").splitlines():
            if not line or line.lstrip().startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            if key.strip() == name:
                return value.strip().strip('"').strip("'")
    return ""


class ApiClient:
    def __init__(self, base_url: str):
        self.base_url = base_url
        self.cookie_jar = http.cookiejar.CookieJar()
        self.opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(self.cookie_jar))

    def request(self, method: str, path: str, payload: dict | None = None) -> tuple[int, bytes, str]:
        body = None
        headers = {}
        if payload is not None:
            body = json.dumps(payload).encode("utf-8")
            headers["Content-Type"] = "application/json"
        request = urllib.request.Request(f"{self.base_url}{path}", data=body, headers=headers, method=method)
        try:
            with self.opener.open(request, timeout=45) as response:
                return response.status, response.read(), response.headers.get("Content-Type", "")
        except urllib.error.HTTPError as exc:
            return exc.code, exc.read(), exc.headers.get("Content-Type", "")

    def json(self, method: str, path: str, payload: dict | None = None, expected_status: int = 200) -> dict:
        status, body, _content_type = self.request(method, path, payload)
        if status != expected_status:
            raise AssertionError(f"{method} {path} returned {status}: {body.decode('utf-8', errors='replace')[:500]}")
        return json.loads(body.decode("utf-8"))


class EndToEndSmokeTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        dart_key = os.environ.get("DART_API_KEY") or read_env_value("DART_API_KEY")
        if not dart_key:
            raise unittest.SkipTest("DART_API_KEY is required for live E2E smoke test.")

        # 포트가 이미 점유돼 있으면 우리가 띄운 서버는 bind에 실패하고, 폴링이 낡은 서버에
        # 붙어 옛 코드를 테스트하게 된다. 서버를 띄우기 전에 포트가 비어 있는지 확인한다.
        with socket.socket() as probe:
            if probe.connect_ex(("127.0.0.1", PORT)) == 0:
                raise RuntimeError(f"Port {PORT} is already in use; stop the stale server before running E2E.")

        env = {
            **os.environ,
            "ADMIN_EMAIL": "admin@example.com",
            "ADMIN_PASSWORD": "test-admin-pass",
            "DART_API_KEY": dart_key,
            "PORT": str(PORT),
        }
        cls.server = subprocess.Popen(
            [sys.executable, "server.py"],  # 테스트와 같은 인터프리터로 부팅 (PATH의 python3에 의존하지 않음)
            cwd=ROOT,
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
        )
        cls.client = ApiClient(BASE_URL)

        deadline = time.time() + 60
        while time.time() < deadline:
            if cls.server.poll() is not None:
                output = cls.server.stdout.read() if cls.server.stdout else ""
                raise RuntimeError(f"Server exited during startup:\n{output}")
            try:
                status, body, _content_type = cls.client.request("GET", "/healthz")
                if status == 200 and json.loads(body).get("status") == "ok":
                    return
            except (OSError, json.JSONDecodeError):
                pass
            time.sleep(1)
        raise TimeoutError("Server did not become ready within 60 seconds.")

    @classmethod
    def tearDownClass(cls):
        server = getattr(cls, "server", None)
        if server and server.poll() is None:
            server.terminate()
            try:
                server.wait(timeout=10)
            except subprocess.TimeoutExpired:
                server.kill()

    def test_dart_to_excel_review_workbook_flow(self):
        status, body, content_type = self.client.request("GET", "/")
        self.assertEqual(status, 200)
        self.assertIn("text/html", content_type)
        self.assertIn(b"root", body)

        login = self.client.json(
            "POST",
            "/api/auth/login",
            {"email": "admin@example.com", "password": "test-admin-pass"},
        )
        self.assertTrue(login["authenticated"])

        project = self.client.json(
            "POST",
            "/api/projects",
            {
                "company_name": "삼성전자",
                "period": "2024",
                "source_standard": "K-GAAP",
                "target_standard": "K-IFRS",
            },
            expected_status=201,
        )
        project_id = project["id"]

        extraction = self.client.json(
            "POST",
            f"/api/projects/{project_id}/dart/import",
            {"corp_code": "00126380", "bsns_year": "2024", "reprt_code": "11011", "fs_div": "CFS"},
            expected_status=201,
        )
        self.assertEqual(extraction["provider"], "dart_api")
        self.assertEqual(extraction["metadata"]["raw_row_count"], 213)
        # 확장 계정 카탈로그(유형자산·이연법인세자산·투자부동산·퇴직급여·영업권 등) 인식과 별칭 사전 확장(134건)으로 변환 대상 증가.
        self.assertEqual(extraction["metadata"]["filtered_row_count"], 25)
        self.assertEqual(len(extraction["rows"]), 25)
        self.assertEqual(
            next(row for row in extraction["rows"] if row["account_name"] == "단기금융상품")["account_key"],
            "financial_instrument",
        )

        accepted = self.client.json(
            "POST",
            f"/api/projects/{project_id}/extractions/{extraction['id']}/accept",
            {},
        )
        self.assertEqual(len(accepted["statements"]), 25)
        financial_product = next(row for row in accepted["statements"] if row["account_name"] == "단기금융상품")
        self.assertEqual(financial_product["standard_code"], "F1000")

        conversion = self.client.json("POST", f"/api/projects/{project_id}/convert", {"responses": {}})
        self.assertEqual(len(conversion["entries"]), 25)
        self.assertIn("단기금융상품", [entry["source_account"] for entry in conversion["entries"]])

        status, workbook_bytes, content_type = self.client.request(
            "GET", f"/api/projects/{project_id}/exports/review-workbook.xlsx"
        )
        self.assertEqual(status, 200)
        self.assertIn("spreadsheetml.sheet", content_type)

        workbook_path = ROOT / "data" / "e2e_review_workbook.xlsx"
        workbook_path.write_bytes(workbook_bytes)
        loaded = load_workbook(workbook_path)
        sheet1, sheet2 = loaded["01_원본_DART"], loaded["02_계정매핑"]
        self.assertEqual(sheet1.max_row, 217)
        self.assertEqual(sheet2.max_row, 26)  # 헤더 1 + 매핑 계정 25행
        sheet1_values = {cell.value for row in sheet1.iter_rows() for cell in row}
        sheet2_values = {cell.value for row in sheet2.iter_rows() for cell in row}
        self.assertIn("자산총계", sheet1_values)
        self.assertIn("제외사유", sheet1_values)
        self.assertIn("단기금융상품", sheet2_values)
        self.assertIn("F1000", sheet2_values)


if __name__ == "__main__":
    unittest.main()
