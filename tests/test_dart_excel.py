import io
import json
import os
import sys
import unittest
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

import gtf_app.dart as dart_module
import server
from gtf_app.dart import dart_raw_statement_rows, dart_statement_rows
from reference_fixture import load_reference

REF = load_reference()


class XlsxUploadRoundTripTests(unittest.TestCase):
    """업로드 xlsx 읽기 경로(openpyxl) 왕복 검증: 쓰기와 같은 라이브러리로 읽는다."""

    def test_parse_xlsx_statement_rows_round_trip(self):
        import tempfile
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["계정명", "금액"])
        sheet.append(["현금및현금성자산", 500000])      # 정수 셀
        sheet.append(["매출채권", 1234.5])              # 실수 셀
        sheet.append(["", ""])                          # 빈 행은 건너뜀
        sheet.append(["충당부채", "1,200,000"])          # 콤마 문자열
        path = Path(tempfile.mkdtemp()) / "upload.xlsx"
        workbook.save(path)

        rows, issues = server.parse_xlsx_statement_rows(path)

        self.assertEqual(issues, [])
        self.assertEqual(
            [(row["account_name"], row["amount"]) for row in rows],
            [("현금및현금성자산", 500000.0), ("매출채권", 1234.5), ("충당부채", 1200000.0)],
        )


class DartImportAndWorkbookTests(unittest.TestCase):
    def test_financial_product_names_do_not_map_to_inventory(self):
        self.assertEqual(server.normalize_account_name("단기금융상품", REF.aliases), "financial_instrument")
        self.assertEqual(server.normalize_account_name("파생상품", REF.aliases), "derivative")  # 전문가 검토 영역

    def test_dart_statement_rows_parse_current_amounts(self):
        payload = {
            "status": "000",
            "message": "정상",
            "list": [
                {
                    "account_nm": "자산총계",
                    "thstrm_amount": "514,531,948,000,000",
                    "sj_nm": "재무상태표",
                },
                {
                    "account_nm": "리스부채",
                    "thstrm_amount": "30,000,000",
                    "sj_nm": "재무상태표",
                    "currency": "KRW",
                    "account_id": "ifrs-full_LeaseLiabilities",
                },
                {
                    "account_nm": "매출",
                    "thstrm_amount": "(250,000,000)",
                    "sj_nm": "손익계산서",
                },
            ],
        }

        rows, issues = dart_statement_rows(payload, REF.aliases)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["account_name"], "리스부채")
        self.assertEqual(rows[0]["amount"], 30_000_000)
        self.assertEqual(rows[0]["account_key"], "lease")
        self.assertEqual(rows[0]["source"], "dart_api")
        self.assertEqual(rows[1]["amount"], -250_000_000)
        self.assertIn("DART 원본 3개 계정 중 변환 대상 핵심 계정 2개", issues[0])

    def test_dart_statement_rows_filters_totals_and_unmapped_accounts(self):
        payload = {
            "status": "000",
            "message": "정상",
            "list": [
                {"account_nm": "자산총계", "thstrm_amount": "1000000", "sj_nm": "재무상태표"},
                {"account_nm": "유동자산", "thstrm_amount": "900000", "sj_nm": "재무상태표"},
                {"account_nm": "현금및현금성자산", "thstrm_amount": "500000", "sj_nm": "재무상태표"},
                {"account_nm": "회원권", "thstrm_amount": "300000", "sj_nm": "재무상태표"},
                {"account_nm": "매출채권", "thstrm_amount": "200000", "sj_nm": "재무상태표"},
            ],
        }

        rows, _issues = dart_statement_rows(payload, REF.aliases)

        self.assertEqual([row["account_name"] for row in rows], ["현금및현금성자산", "매출채권"])

    def test_dart_raw_statement_rows_preserves_filtered_out_accounts(self):
        payload = {
            "status": "000",
            "message": "정상",
            "list": [
                {"account_nm": "자산총계", "thstrm_amount": "1000000", "sj_nm": "재무상태표"},
                {"account_nm": "현금및현금성자산", "thstrm_amount": "500000", "sj_nm": "재무상태표"},
            ],
        }

        raw_rows = dart_raw_statement_rows(payload, REF.aliases)
        filtered_rows, _issues = dart_statement_rows(payload, REF.aliases)

        self.assertEqual([row["account_name"] for row in raw_rows], ["자산총계", "현금및현금성자산"])
        self.assertEqual([row["account_name"] for row in filtered_rows], ["현금및현금성자산"])
        self.assertFalse(raw_rows[0]["conversion_candidate"])
        self.assertIn("합계", raw_rows[0]["filter_reason"])

    def test_dart_raw_rows_from_upload_reads_saved_payload(self):
        saved_payload = {
            "raw_rows": [{"account_name": "자산총계", "conversion_candidate": False}],
            "filtered_rows": [{"account_name": "현금및현금성자산"}],
        }

        rows = server.dart_raw_rows_from_upload({"file_bytes": json.dumps(saved_payload).encode("utf-8")})

        self.assertEqual(rows, saved_payload["raw_rows"])

    def test_dart_statement_rows_reports_api_error(self):
        rows, issues = dart_statement_rows({"status": "013", "message": "조회된 데이터가 없습니다."}, REF.aliases)

        self.assertEqual(rows, [])
        self.assertIn("DART API 오류 013", issues[0])

    def test_review_workbook_contains_expected_sheets(self):
        workbook = server.review_workbook_bytes(
            {"company_name": "샘플테크", "period": "2024"},
            [
                {
                    "account_name": "자산총계",
                    "amount": 100_000_000,
                    "statement_type": "재무상태표",
                    "conversion_candidate": False,
                    "filter_reason": "합계/소계/성과지표/현금흐름표 활동 항목",
                },
                {"account_name": "리스부채", "amount": 30_000_000, "statement_type": "재무상태표"},
            ],
            [
                {
                    "account_name": "리스부채",
                    "normalized_account": "리스",
                    "standard_code": "A2100",
                    "amount": 30_000_000,
                    "period": "2024",
                    "mapping_type": "judgment",
                    "rule_summary": "K-IFRS 제1116호 검토",
                }
            ],
            {
                "entries": [
                    {
                        "source_account": "리스부채",
                        "standard_code": "A2100",
                        "target_account": "사용권자산",
                        "statement_type": "재무상태표",
                        "statement_line_item": "사용권자산",
                        "amount": 30_000_000,
                        "adjustment": 3_812_000,
                        "mapping_type": "judgment",
                        "calculation": "리스료 현재가치",
                    }
                ],
                "draft_notes": [{"account": "리스부채", "draft_note": "K-IFRS 제1116호 검토 필요"}],
                "ai_assistance": {"status": "skipped", "overall_note": "사람 검토 필요"},
            },
            [{"created_at": "2026-07-03T00:00:00+00:00", "actor": "system", "event_type": "test", "detail": {}}],
        )

        loaded = load_workbook(io.BytesIO(workbook))
        self.assertEqual(len(loaded.sheetnames), 6)
        self.assertIn("01_원본_DART", loaded.sheetnames)
        self.assertIn("05_전환조정요약", loaded.sheetnames)
        self.assertIn("06_감사로그", loaded.sheetnames)
        sheet1_values = {cell.value for row in loaded["01_원본_DART"].iter_rows() for cell in row}
        self.assertIn("자산총계", sheet1_values)
        self.assertIn("제외사유", sheet1_values)

    def test_transition_summary_aggregates_by_section_sign(self):
        # 전환조정 요약(1101호 문단 10): 자산(+)·부채(−) 부호를 반영한 순자산 영향 집계.
        from gtf_app.excel_export import transition_summary_rows

        entries = [
            {"standard_code": "A2100", "adjustment": 1_000},   # 사용권자산 +1,000
            {"standard_code": "L2150", "adjustment": 1_200},   # 리스부채 +1,200 (순자산 −1,200)
            {"standard_code": "R4000", "adjustment": 50},      # 차입원가 자본화 +50
            {"standard_code": "E1050", "adjustment": 0},       # 재분류 — 집계 제외
        ]
        rows = transition_summary_rows(entries)
        by_label = {row[0]: row for row in rows if row and len(row) >= 4}
        self.assertEqual(by_label["자산 조정"][3], 1_000)
        self.assertEqual(by_label["부채 조정"][3], -1_200)
        self.assertEqual(by_label["손익 조정"][3], 50)
        self.assertEqual(by_label["순자산(자본총계) 영향 합계"][3], -150)
        self.assertEqual(by_label["자본 조정"][1], 0)  # 조정액 0 재분류는 건수에서 제외

    def test_routes_include_dart_import_and_workbook_export(self):
        routes = {(sorted(r.methods)[0], r.path) for r in server.app.routes if hasattr(r, "methods")}
        self.assertIn(("POST", "/api/projects/{project_id}/dart/import"), routes)
        self.assertIn(("POST", "/api/projects/{project_id}/dart/reports"), routes)
        self.assertIn(("GET", "/api/projects/{project_id}/exports/{export_name}"), routes)
        self.assertIn(("DELETE", "/api/projects/{project_id}"), routes)
        self.assertIn(("POST", "/api/projects/{project_id}/policy-comparison"), routes)

    def test_dart_available_reports_filters_periodic_reports(self):
        original_request = dart_module.dart_json_request
        original_key = server.os.environ.get("DART_API_KEY")

        def fake_request(_url, params, timeout=30):
            self.assertEqual(params["corp_code"], "00126380")
            return {
                "status": "000",
                "message": "정상",
                "list": [
                    {
                        "corp_name": "삼성전자",
                        "report_nm": "분기보고서 (2026.03)",
                        "rcept_no": "20260515000001",
                        "rcept_dt": "20260515",
                    },
                    {
                        "corp_name": "삼성전자",
                        "report_nm": "사업보고서 (2025.12)",
                        "rcept_no": "20260320000001",
                        "rcept_dt": "20260320",
                    },
                    {
                        "corp_name": "삼성전자",
                        "report_nm": "주요사항보고서",
                        "rcept_no": "20260401000001",
                        "rcept_dt": "20260401",
                    },
                ],
            }

        try:
            server.os.environ["DART_API_KEY"] = "test-key"
            dart_module.dart_json_request = fake_request
            reports, issues, metadata = server.fetch_dart_available_reports({"corp_code": "00126380", "from_year": "2026", "to_year": "2026"})
        finally:
            dart_module.dart_json_request = original_request
            if original_key is None:
                server.os.environ.pop("DART_API_KEY", None)
            else:
                server.os.environ["DART_API_KEY"] = original_key

        self.assertEqual(issues, [])
        self.assertEqual(metadata["corp_code"], "00126380")
        self.assertEqual(len(reports), 2)
        self.assertEqual(reports[0]["bsns_year"], "2026")
        self.assertEqual(reports[0]["reprt_code"], "11013")
        self.assertEqual(reports[1]["bsns_year"], "2025")
        self.assertEqual(reports[1]["reprt_code"], "11011")


if __name__ == "__main__":
    unittest.main()
