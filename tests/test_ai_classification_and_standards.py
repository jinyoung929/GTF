"""
AI 1차 분류 및 기준서 문단 검색 DB 테스트
==========================================

검증 범위:
  - standards_paragraphs 시드: K-GAAP/K-IFRS 분리 보관
  - find_standards_paragraphs: 계정키/기준세트/키워드 검색
  - generate_conversion: 판단 필요 항목에 기준서 문단 첨부
  - conversion_basis_report / 엑셀 워크북: 문단 근거 표시
  - call_ai_classification: 키 미설정/미분류 없음/정상 응답/무효 제안 방어
  - build_statement_record: AI 제안의 사람 확정 적용 (X9999 한정)

실행:
    python3 -m unittest tests.test_ai_classification_and_standards -v
"""

import io
import json
import os
import sys
import unittest
from unittest import mock

_HERE = os.path.dirname(os.path.abspath(__file__))
for _candidate in (_HERE, os.path.dirname(_HERE)):
    if os.path.exists(os.path.join(_candidate, "server.py")):
        sys.path.insert(0, _candidate)
        break

from openpyxl import load_workbook  # noqa: E402

import server  # noqa: E402
from gtf_app.domain import (  # noqa: E402
    build_statement_record,
    conversion_basis_report,
    generate_conversion,
)
from gtf_app.excel_export import review_workbook_bytes  # noqa: E402

sys.path.insert(0, _HERE)
from reference_fixture import load_reference, paragraph_session  # noqa: E402

# 기준정보는 운영과 동일하게 SQL 시드 → DB 조회로 로드해 주입한다
REF = load_reference()


PROJECT = {
    "id": "proj-test",
    "company_name": "테스트회사",
    "period": "2024",
    "source_standard": "K-GAAP",
    "target_standard": "IFRS",
}


class StandardsParagraphsSeedTest(unittest.TestCase):
    def test_both_standard_sets_are_seeded_separately(self):
        conn = paragraph_session()
        rows = server.find_standards_paragraphs(conn)
        self.assertTrue(rows)
        sets = {row["standard_set"] for row in rows}
        self.assertEqual(sets, {"K-GAAP", "K-IFRS"})

    def test_seed_is_idempotent(self):
        conn = paragraph_session()
        before = len(server.find_standards_paragraphs(conn))
        server.ensure_standards_paragraphs(conn)
        self.assertEqual(len(server.find_standards_paragraphs(conn)), before)

    def test_judgment_account_has_paragraphs_from_both_sets(self):
        conn = paragraph_session()
        for account_key in ("lease", "development", "revenue", "financial_instrument", "provision", "receivables"):
            sets = {row["standard_set"] for row in server.find_standards_paragraphs(conn, account_key=account_key)}
            self.assertEqual(sets, {"K-GAAP", "K-IFRS"}, account_key)


class StandardsParagraphsSearchTest(unittest.TestCase):
    def test_keyword_search_hits_content_and_keywords(self):
        conn = paragraph_session()
        hits = server.find_standards_paragraphs(conn, query="기대신용손실")
        self.assertTrue(hits)
        self.assertTrue(all("receivables" == row["account_key"] for row in hits))

    def test_standard_set_filter(self):
        conn = paragraph_session()
        hits = server.find_standards_paragraphs(conn, account_key="lease", standard_set="K-GAAP")
        self.assertEqual({row["standard_set"] for row in hits}, {"K-GAAP"})
        self.assertIn("운용리스", hits[0]["content"])

    def test_no_match_returns_empty(self):
        conn = paragraph_session()
        self.assertEqual(server.find_standards_paragraphs(conn, query="존재하지않는키워드"), [])


class ConversionParagraphAttachmentTest(unittest.TestCase):
    def make_statement(self, account_name, mapping_type="judgment"):
        return {
            "id": "row-1",
            "account_name": account_name,
            "standard_code": "A2100",
            "amount": 1000.0,
            "mapping_type": mapping_type,
            "rule_summary": "테스트 규칙",
        }

    def test_judgment_item_includes_both_standard_sets(self):
        output = generate_conversion(PROJECT, [self.make_statement("리스부채")], {}, REF)
        paragraphs = output["judgment_items"][0]["standards_paragraphs"]
        self.assertTrue(paragraphs)
        self.assertEqual({p["standard_set"] for p in paragraphs}, {"K-GAAP", "K-IFRS"})

    def test_basis_report_lists_paragraphs(self):
        output = generate_conversion(PROJECT, [self.make_statement("리스부채")], {}, REF)
        report = conversion_basis_report(output)
        self.assertIn("K-IFRS 제1116호", report)
        self.assertIn("일반기업회계기준 제13장", report)

    def test_excel_workbook_review_sheet_contains_paragraphs(self):
        output = generate_conversion(PROJECT, [self.make_statement("리스부채")], {}, REF)
        workbook = review_workbook_bytes(PROJECT, [], [], output, [])
        sheet4 = load_workbook(io.BytesIO(workbook))["04_KIFRS_검토근거"]
        text = " ".join(str(cell.value) for row in sheet4.iter_rows() for cell in row if cell.value)
        self.assertIn("기준서 문단", text)
        self.assertIn("일반기업회계기준", text)


class AiClassificationCallTest(unittest.TestCase):
    def test_no_unmapped_accounts_skips(self):
        result = server.call_ai_classification([])
        self.assertEqual(result["status"], "skipped")
        self.assertEqual(result["suggestions"], {})

    def test_missing_api_key_reports_not_configured(self):
        with mock.patch.dict(os.environ, {"OPENAI_API_KEY": ""}):
            result = server.call_ai_classification(["임차보증금"])
        self.assertEqual(result["status"], "not_configured")
        self.assertTrue(result["human_review_required"])

    def _fake_openai(self, items):
        # openai SDK 클라이언트 흉내: responses.create는 output_text 응답, embeddings.create는
        # 빈 결과(→ openai_embed가 None을 돌려 키워드 폴백 경로를 타게 함)를 반환한다.
        client = mock.Mock()
        client.responses.create.return_value = mock.Mock(
            output_text=json.dumps({"items": items}, ensure_ascii=False)
        )
        client.embeddings.create.return_value = mock.Mock(data=[])
        return client

    def test_connected_response_builds_validated_suggestions(self):
        items = [
            {
                "account_name": "임차보증금",
                "suggested_account_key": "financial_instrument",
                "confidence": "medium",
                "rationale": "임차보증금은 계약 종료 시 반환받을 계약상 권리다. K-IFRS 제1109호상 금융자산의 정의를 충족한다. 따라서 금융상품으로 분류하는 것이 적합하다.",
                "basis_reference": "K-IFRS 제1109호 금융상품",
                "alternative_account_key": "deposits",
                "alternative_rejected_reason": "장기 예치 성격보다 계약상 반환 권리가 본질이므로 배제했다.",
            },
            {"account_name": "이상한계정", "suggested_account_key": "없는키", "confidence": "low", "rationale": "-",
             "basis_reference": "", "alternative_account_key": "", "alternative_rejected_reason": ""},
            {"account_name": "모르는계정", "suggested_account_key": "other", "confidence": "low", "rationale": "-",
             "basis_reference": "", "alternative_account_key": "", "alternative_rejected_reason": ""},
        ]
        session = paragraph_session()
        with mock.patch.dict(os.environ, {"OPENAI_API_KEY": "sk-test"}):
            with mock.patch.object(server, "OpenAI", return_value=self._fake_openai(items)):
                result = server.call_ai_classification(["임차보증금", "이상한계정", "모르는계정"], session)
        self.assertEqual(result["status"], "connected")
        self.assertEqual(set(result["suggestions"].keys()), {"임차보증금"})
        suggestion = result["suggestions"]["임차보증금"]
        self.assertEqual(suggestion["account_key"], "financial_instrument")
        self.assertEqual(suggestion["basis_reference"], "K-IFRS 제1109호 금융상품")
        self.assertEqual(suggestion["alternative_label"], "보증금")  # 계정키 → 화면용 라벨 변환
        self.assertIn("배제", suggestion["alternative_rejected_reason"])
        self.assertTrue(suggestion["human_review_required"])

    def test_prompt_includes_retrieved_standards_for_grounding(self):
        # 분류 프롬프트에 계정별 기준서 문단(RAG)이 첨부되는지 — 근거 접지의 핵심 계약
        captured = {}

        def capture_create(**kwargs):
            captured.update(kwargs)
            return mock.Mock(output_text='{"items": []}')

        client = mock.Mock()
        client.responses.create.side_effect = capture_create
        client.embeddings.create.return_value = mock.Mock(data=[])
        session = paragraph_session()
        with mock.patch.dict(os.environ, {"OPENAI_API_KEY": "sk-test"}):
            with mock.patch.object(server, "OpenAI", return_value=client):
                server.call_ai_classification(["리스보증금"], session)
        prompt_text = captured["input"][0]["content"][0]["text"]
        self.assertIn("retrieved_standards", prompt_text)
        self.assertIn("reference_code", captured["instructions"])  # 인용 강제 문구
        self.assertNotIn("근거를 한국어 한 문장으로", captured["instructions"])  # 옛 한-문장 제약 제거 확인

    def test_attach_only_marks_unmapped_rows(self):
        rows = [
            {"account_name": "임차보증금", "amount": 500.0},
            {"account_name": "현금", "amount": 100.0},
        ]
        items = [
            {
                "account_name": "임차보증금",
                "suggested_account_key": "financial_instrument",
                "confidence": "medium",
                "rationale": "보증금은 금융자산 성격입니다.",
            }
        ]
        with mock.patch.dict(os.environ, {"OPENAI_API_KEY": "sk-test"}):
            with mock.patch.object(server, "OpenAI", return_value=self._fake_openai(items)):
                rows, result = server.attach_ai_classification(rows, paragraph_session())
        self.assertEqual(result["status"], "connected")
        self.assertIn("ai_suggestion", rows[0])
        self.assertNotIn("ai_suggestion", rows[1])


class ExpandedStandardCatalogTest(unittest.TestCase):
    def test_catalog_includes_ifrs_accounts_beyond_core_eight(self):
        for key in ("ppe", "deferred_tax_asset", "prepaid_expense", "share_capital", "retained_earnings", "cost_of_sales"):
            self.assertIn(key, REF.accounts)

    def test_ai_can_suggest_expanded_account(self):
        # 임차보증금은 키워드 사전에 없어 미분류(X9999)로 빠지므로 AI 제안 대상이 된다.
        row = {
            "account_name": "임차보증금",
            "amount": 60000000,
            "ai_suggestion": {"account_key": "deposits", "label": "보증금", "confidence": "high", "rationale": "반환 예정 보증금은 금융자산 성격"},
        }
        record = build_statement_record("2024", row, REF)
        self.assertEqual(record["standard_code"], "A1400")
        self.assertEqual(record["mapping_source"], "ai_suggested_human_accepted")

    def test_classification_candidates_exclude_only_other(self):
        candidates = [key for key in REF.accounts if key != "other"]
        self.assertGreater(len(candidates), 8)
        self.assertNotIn("other", candidates)

    def test_seed_reference_data_handles_expanded_accounts(self):
        # 보조 조회 테이블(ifrs_accounts 등)은 핵심 시드 테이블에서 파생 생성된다
        from sqlalchemy import func, select

        from gtf_app.models import IfrsAccount, StandardAccount
        from reference_fixture import seeded_session

        session = seeded_session()
        server.seed_reference_data(session)
        accounts = session.scalar(select(func.count()).select_from(StandardAccount))
        derived = session.scalar(select(func.count()).select_from(IfrsAccount))
        self.assertEqual(derived, accounts)
        self.assertEqual(accounts, len(REF.accounts))


class StatementTemplateSeedTest(unittest.TestCase):
    """표준양식 라인 SQL 시드(단일 출처)가 전 계정을 커버하고 코드 도출 순서와 일치하는지 검증."""

    def setUp(self):
        from reference_fixture import seeded_session

        self.conn = seeded_session()

    def tearDown(self):
        self.conn.close()

    def rows(self):
        from sqlalchemy import select

        from gtf_app.models import FinancialStatementTemplate

        return self.conn.execute(
            select(
                FinancialStatementTemplate.account_key,
                FinancialStatementTemplate.display_order,
                FinancialStatementTemplate.line_item,
            )
        ).all()

    def test_every_standard_account_has_a_template_line(self):
        mapped = {row.account_key for row in self.rows()}
        self.assertEqual(mapped, set(REF.accounts.keys()))

    def test_display_order_matches_account_code_derivation(self):
        from gtf_app.domain import account_presentation_order

        for row in self.rows():
            code = REF.accounts[row.account_key]["code"]
            self.assertEqual(
                row.display_order,
                account_presentation_order(code),
                f"{row.account_key}의 SQL display_order가 계정코드 도출값과 다릅니다",
            )

    def test_seed_is_idempotent(self):
        server.ensure_reference_accounts(self.conn)
        server.ensure_statement_templates(self.conn)
        self.assertEqual(len(self.rows()), len(REF.accounts))

    def test_new_catalog_accounts_get_statement_lines_in_conversion(self):
        statement = build_statement_record(
            "2024",
            {
                "account_name": "이연법인세자산",
                "amount": 35000000.0,
                "ai_suggestion": {"account_key": "deferred_tax_asset", "label": "이연법인세자산", "confidence": "high", "rationale": "IAS 12"},
            },
            REF,
        )
        output = generate_conversion(PROJECT, [statement], {}, REF)
        entry = output["entries"][0]
        self.assertEqual(entry["statement_line_item"], "이연법인세자산")
        self.assertEqual(entry["statement_type"], "재무상태표")


class PresentationOrderTest(unittest.TestCase):
    def test_order_derived_from_account_code(self):
        from gtf_app.domain import account_presentation_order

        # 자산(A) < 금융상품(F) < 부채(L) < 자본(E) < 손익(R) < 미분류(X)
        self.assertLess(account_presentation_order("A1000"), account_presentation_order("A1100"))
        self.assertLess(account_presentation_order("A3100"), account_presentation_order("F1000"))
        self.assertLess(account_presentation_order("F1000"), account_presentation_order("L1000"))
        self.assertLess(account_presentation_order("L2200"), account_presentation_order("E1000"))
        self.assertLess(account_presentation_order("E1200"), account_presentation_order("R1000"))
        self.assertLess(account_presentation_order("R3000"), account_presentation_order("X9999"))

    def test_malformed_code_does_not_crash(self):
        from gtf_app.domain import account_presentation_order

        self.assertIsInstance(account_presentation_order(""), int)
        self.assertIsInstance(account_presentation_order("ZZZZ"), int)

    def test_conversion_entries_sorted_by_code(self):
        stmts = [
            build_statement_record("2024", {"account_name": "매출액", "amount": 1000}, REF),
            build_statement_record("2024", {"account_name": "현금및현금성자산", "amount": 1000}, REF),
            build_statement_record("2024", {"account_name": "충당부채", "amount": 1000}, REF),
            build_statement_record("2024", {"account_name": "매출채권", "amount": 1000}, REF),
        ]
        output = generate_conversion(PROJECT, stmts, {}, REF)
        codes = [entry["standard_code"] for entry in output["entries"]]
        self.assertEqual(codes, sorted(codes, key=lambda c: __import__("server").account_presentation_order(c)))
        self.assertEqual(codes[0], "A1000")  # 현금이 맨 앞


class AiSuggestionHumanAcceptanceTest(unittest.TestCase):
    SUGGESTION = {
        "account_key": "financial_instrument",
        "label": "금융상품",
        "confidence": "medium",
        "rationale": "보증금은 금융자산 성격입니다.",
        "provider": "openai",
        "model": "gpt-4.1-mini",
        "human_review_required": True,
    }

    def test_suggestion_applied_only_for_unmapped_account(self):
        record = build_statement_record("2024", {"account_name": "임차보증금", "amount": 500.0, "ai_suggestion": self.SUGGESTION}, REF)
        self.assertEqual(record["standard_code"], "F1000")
        self.assertEqual(record["mapping_source"], "ai_suggested_human_accepted")
        self.assertIn("[AI 1차 분류 제안", record["rule_summary"])

    def test_suggestion_ignored_when_rule_mapping_succeeds(self):
        record = build_statement_record("2024", {"account_name": "리스부채", "amount": 500.0, "ai_suggestion": self.SUGGESTION}, REF)
        self.assertEqual(record["standard_code"], "A2100")
        self.assertEqual(record["mapping_source"], "rule_based")

    def test_invalid_suggestion_key_falls_back_to_manual_review(self):
        record = build_statement_record("2024", {"account_name": "임차보증금", "amount": 500.0, "ai_suggestion": {"account_key": "없는키"}}, REF)
        self.assertEqual(record["standard_code"], "X9999")
        self.assertEqual(record["mapping_source"], "rule_based")

    def test_plain_record_keeps_previous_shape(self):
        record = build_statement_record("2024", {"account_name": "현금", "amount": 1.0}, REF)
        self.assertEqual(record["standard_code"], "A1000")
        self.assertEqual(record["mapping_source"], "rule_based")
        self.assertIsNone(record["ai_suggestion"])


if __name__ == "__main__":
    unittest.main()
