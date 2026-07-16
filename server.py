"""GTF 서버 — K-GAAP → K-IFRS 변환 검토 서비스의 FastAPI 앱.

이 파일은 HTTP·DB·외부 API 등 부수효과가 있는 코드의 단일 진입점이다.
순수 회계 로직은 gtf_app/domain.py, DART 연동은 gtf_app/dart.py,
ORM 모델은 gtf_app/models.py, 엔진 설정은 gtf_app/db.py에 있다.

구성 (위에서 아래로):
  §1  경로·상수
  §2  환경·설정, 직렬화 헬퍼
  §3  DB 엔진·세션 (configure_engine / get_db)
  §4  부팅: 기준정보 캐시·스키마 드리프트 치유·init_db
  §5  시드·기준정보 로더 (seeds/*.sql → ReferenceData)
  §6  감사로그·행 직렬화
  §7  파일 파서 (xlsx / pdf)
  §8  외부 서비스 설정 (OCR·AI·DART)
  §9  AI·RAG (임베딩 검색, 판단보조, 1차 분류, Gemini OCR)
  §10 FastAPI 앱: 의존성 → 요청 모델 → 라우트 → 정적 서빙 → main
"""

from __future__ import annotations

import base64
import contextlib
import importlib.util
import json
import math
import mimetypes
import os
import re
import secrets
import uuid
import zipfile
from decimal import Decimal
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator

import openai
import requests
import uvicorn
from fastapi import Body, Depends, FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, PlainTextResponse, Response
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel
from sqlalchemy import DateTime, String, delete, func, inspect, select, text, update
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session, sessionmaker

from gtf_app.db import backend, create_db_engine, create_session_factory
from gtf_app.models import (
    AppUser,
    AuditLog,
    Base,
    ChecklistItem,
    Conversion,
    Extraction,
    FinancialStatementTemplate,
    KgaapAccount,
    Project,
    Review,
    StandardAccount,
    StandardsParagraph,
    Statement,
    Upload,
    UserSession,
)

from gtf_app.auth import (
    admin_config,
    hash_password,
    normalize_email,
    session_token_hash,
    user_public_dict,
    verify_password,
)
from gtf_app.dart import (
    dart_raw_rows_from_upload,
    fetch_dart_available_reports,
    fetch_dart_statement_rows,
)
from gtf_app.domain import (
    ReferenceData,
    account_presentation_order,
    build_review_summary,
    build_statement_record,
    utc_now,
    conversion_adjustments_csv,
    compare_policy_scenarios,
    conversion_basis_report,
    generate_conversion,
    looks_numeric,
    normalize_account_name,
    parse_amount,
    parse_statement_rows,
    verify_reference_contract,
    validate_statement_records,
)
from gtf_app.excel_export import review_workbook_bytes


# ───────────────────────────────────────────────────────────────────────────
# §1 경로·상수
# ───────────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
DB_PATH = DATA_DIR / "gtf.sqlite3"
SEED_DIR = ROOT / "seeds"
FIGMA_DIST_DIR = ROOT / "figma_make" / "dist"
ENV_PATHS = (ROOT / ".env", ROOT / ".env.local")
GEMINI_INLINE_LIMIT_BYTES = 20 * 1024 * 1024
OPENAI_DEFAULT_MODEL = "gpt-4.1-mini"
EMBEDDING_MODEL = os.environ.get("OPENAI_EMBEDDING_MODEL", "text-embedding-3-small").strip() or "text-embedding-3-small"
SESSION_COOKIE = "gtf_session"
SESSION_MAX_AGE_SECONDS = 60 * 60 * 24 * 7
ADMIN_USER_ID = "admin"
DEMO_USER_ID = "demo"
DEFAULT_DEMO_EMAIL = "demo@gtf.local"
DEFAULT_DEMO_PASSWORD = "change-this-demo-password"
INDEX_HTML = """<!doctype html>
<html lang="ko">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>GTF</title></head>
<body><main><h1>GTF</h1><p>Figma UI build is not available. Run the frontend build and try again.</p></main></body>
</html>"""
STYLES_CSS = "body{font-family:system-ui,sans-serif;margin:2rem}"
APP_JS = "console.info('GTF fallback bundle')"


# ───────────────────────────────────────────────────────────────────────────
# §2 환경·설정, 직렬화 헬퍼
# ───────────────────────────────────────────────────────────────────────────
def load_local_env() -> None:
    for env_path in ENV_PATHS:
        if not env_path.exists():
            continue
        for raw_line in env_path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


load_local_env()


def database_config() -> dict:
    """/healthz가 노출하는 DB 설정 요약 (실제 접속은 gtf_app/db.py의 엔진이 담당한다)."""
    database_url = os.environ.get("DATABASE_URL") or os.environ.get("NEON_DATABASE_URL") or ""
    postgres_driver_ready = importlib.util.find_spec("psycopg") is not None
    current_backend = backend()
    return {
        "backend": current_backend,
        "sqlite_path": str(DB_PATH),
        "sqlite_ready": DB_PATH.exists(),
        "database_url_ready": bool(database_url),
        "postgres_driver_ready": postgres_driver_ready,
        "postgres_ready": current_backend == "postgres" and bool(database_url) and postgres_driver_ready,
    }


def normalize_db_value(value):
    if isinstance(value, uuid.UUID):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, list):
        return [normalize_db_value(item) for item in value]
    if isinstance(value, dict):
        return {key: normalize_db_value(item) for key, item in value.items()}
    return value


def parse_json_field(value, fallback):
    if value is None:
        return fallback
    if isinstance(value, (list, dict)):
        return normalize_db_value(value)
    if isinstance(value, str):
        return json.loads(value)
    return value


def upload_public_dict(upload_or_dict) -> dict:
    """업로드 응답에서 원본 바이트를 빼고, 파일이 DB에 있는지 여부만 노출한다."""
    upload = dict(upload_or_dict) if isinstance(upload_or_dict, dict) else row_to_dict(upload_or_dict)
    upload["db_file_ready"] = bool(upload.pop("file_bytes", None))
    return upload


def figma_static_file(path: str) -> Path | None:
    if not FIGMA_DIST_DIR.exists():
        return None
    relative = "index.html" if path == "/" else path.lstrip("/")
    candidate = (FIGMA_DIST_DIR / relative).resolve()
    try:
        candidate.relative_to(FIGMA_DIST_DIR.resolve())
    except ValueError:
        return None
    return candidate if candidate.is_file() else None


engine: Engine | None = None
SessionLocal: sessionmaker[Session] | None = None


# ───────────────────────────────────────────────────────────────────────────
# §3 DB 엔진·세션
# ───────────────────────────────────────────────────────────────────────────
def configure_engine() -> Engine:
    """현재 DB_PATH/환경변수로 엔진과 세션 팩토리를 (재)생성한다.

    테스트가 server.DB_PATH를 임시 경로로 바꾼 뒤 init_db()를 부르므로 재생성 가능해야 한다.
    """
    global engine, SessionLocal
    engine = create_db_engine(DB_PATH)
    SessionLocal = create_session_factory(engine)
    return engine


def get_session() -> Session:
    """세션 하나를 연다. 스크립트·시드 등 FastAPI 밖에서 쓴다."""
    if SessionLocal is None:
        configure_engine()
    return SessionLocal()


def get_db() -> Iterator[Session]:
    """FastAPI 의존성: 요청마다 세션을 열고 끝나면 닫는다."""
    with get_session() as session:
        yield session


# ───────────────────────────────────────────────────────────────────────────
# §4 부팅: 기준정보 캐시·스키마 치유·init_db
# ───────────────────────────────────────────────────────────────────────────
# 기준정보(계정·별칭·체크리스트·양식·문단)를 DB에서 한 번 로드해 두는 서버 캐시.
# 시드 이후에만 바뀌므로(런타임 변경 없음) 매 요청 재조회 대신 이 캐시를 읽는다.
# init_db 끝에서 refresh_reference_cache가 채우고, 계약 검증까지 통과해야 서버가 뜬다.
REFERENCE = ReferenceData()


def refresh_reference_cache(session: Session) -> ReferenceData:
    """DB 기준정보를 로드해 계약을 검증하고 전역 REFERENCE 캐시를 갱신한다.

    계산기가 요구하는 계정키·체크리스트 키가 SQL 시드에 빠져 있으면(조정=0 무증상 버그)
    여기서 RuntimeError로 서버 시작을 실패시킨다.
    """
    global REFERENCE
    ref = load_reference_data(session)
    errors = verify_reference_contract(ref)
    if errors:
        raise RuntimeError(
            "기준정보 계약 검증 실패 (SQL 시드와 계산기 코드가 어긋남):\n  - " + "\n  - ".join(errors)
        )
    REFERENCE = ref
    return ref


def resolve_postgres_type_drift(session: Session) -> None:
    """옛 수동 스키마가 만든 uuid/timestamptz/jsonb 컬럼을 ORM 타입(text)으로 수렴시킨다.

    create_all은 기존 테이블을 절대 변경하지 않는다. 그래서 과거에 postgres/schema.sql을
    손으로 적용한 배포 DB에서는 컬럼 타입이 ORM(text)과 어긋나, 첫 UPDATE/INSERT가
    DatatypeMismatch(f405)로 죽고 서버가 못 뜬다. 여기서 어긋난 컬럼만 골라
    ALTER TYPE text USING ...으로 값을 보존한 채 변환한다
    (uuid → 같은 표기의 문자열, jsonb → JSON 문자열, timestamptz → ISO8601 문자열).
    uuid PK/FK는 타입을 바꾸면 uuid=text 비교가 불가능해 FK 제약을 먼저 제거한다
    (앱은 DB FK에 의존하지 않고 자체적으로 참조 무결성을 관리한다).
    """
    if backend() != "postgres":
        return
    inspector = inspect(session.bind)
    drifted: dict[str, list[tuple[str, bool]]] = {}
    for table in Base.metadata.sorted_tables:
        if not inspector.has_table(table.name):
            continue
        live_types = {c["name"]: c["type"] for c in inspector.get_columns(table.name)}
        for column in table.columns:
            live = live_types.get(column.name)
            if live is None or not isinstance(column.type, String) or isinstance(live, String):
                continue
            drifted.setdefault(table.name, []).append((column.name, isinstance(live, DateTime)))
    if not drifted:
        return

    for table in Base.metadata.sorted_tables:
        if not inspector.has_table(table.name):
            continue
        for fk in inspector.get_foreign_keys(table.name):
            if (table.name in drifted or fk.get("referred_table") in drifted) and fk.get("name"):
                session.execute(text(f'ALTER TABLE {table.name} DROP CONSTRAINT IF EXISTS "{fk["name"]}"'))

    for table_name, columns in drifted.items():
        for column_name, is_timestamp in columns:
            using = (
                f"to_char({column_name} AT TIME ZONE 'UTC', 'YYYY-MM-DD\"T\"HH24:MI:SS.US') || '+00:00'"
                if is_timestamp
                else f"{column_name}::text"
            )
            session.execute(text(f"ALTER TABLE {table_name} ALTER COLUMN {column_name} DROP DEFAULT"))
            session.execute(text(f"ALTER TABLE {table_name} ALTER COLUMN {column_name} TYPE text USING {using}"))
    session.commit()


def migrate_legacy_columns(session: Session) -> None:
    """기존 배포 DB에 없던 컬럼을 추가한다 (새 DB는 create_all이 이미 만들었다)."""
    existing = {column["name"] for column in inspect(session.bind).get_columns("standards_paragraphs")}
    is_postgres = backend() == "postgres"
    if "embedding" not in existing:
        session.execute(text("ALTER TABLE standards_paragraphs ADD COLUMN embedding text"))
    if "content_hash" not in existing:
        session.execute(text("ALTER TABLE standards_paragraphs ADD COLUMN content_hash text"))
    alias_columns = {column["name"] for column in inspect(session.bind).get_columns("kgaap_accounts")}
    if "match_priority" in alias_columns:
        # 별칭 길이는 별칭 자체에서 파생되므로 컬럼으로 이중 저장할 이유가 없다 (단순화 감사로 은퇴).
        session.execute(text("ALTER TABLE kgaap_accounts DROP COLUMN match_priority"))
    checklist_columns = {column["name"] for column in inspect(session.bind).get_columns("checklist_items")}
    if "options" not in checklist_columns:
        session.execute(text("ALTER TABLE checklist_items ADD COLUMN options text NOT NULL DEFAULT ''"))
    if is_postgres:
        # 옛 수동 스키마 시절의 CHECK 제약 잔재 제거. ORM 모델은 CHECK를 정의하지 않으므로
        # 우리 테이블에 남은 CHECK는 전부 레거시이며, 값 집합이 늘어날 때 시드를 거부한다
        # (실사례: input_type CHECK가 새 'choice' 값을 거부해 배포 부팅 실패).
        legacy_checks = session.execute(
            text(
                "SELECT conrelid::regclass::text AS table_name, conname FROM pg_constraint "
                "WHERE contype = 'c' AND conrelid != 0 AND connamespace = 'public'::regnamespace"
            )
        ).all()
        for table_name, constraint_name in legacy_checks:
            if table_name in Base.metadata.tables:
                session.execute(text(f'ALTER TABLE {table_name} DROP CONSTRAINT "{constraint_name}"'))

        # 시드 소유 테이블(시드가 전량 재생성)은 ORM 모델이 구조의 단일 출처다. 모델에 없는
        # 여분 컬럼(옛 수동 스키마 잔재)이 NOT NULL이면 시드 INSERT가 즉사하고, 레거시 UNIQUE
        # 제약은 upsert를 거부할 수 있으므로 함께 정리한다. 데이터 테이블(projects 등)은
        # 사용자 데이터가 있으므로 건드리지 않는다.
        seed_owned = {
            "standard_accounts", "kgaap_accounts", "checklist_items",
            "financial_statement_templates", "standards_paragraphs",
        }
        derived_columns = {"standards_paragraphs": {"embedding_vec"}}  # pgvector 파생 컬럼은 보존
        for table_name in seed_owned:
            model_columns = set(Base.metadata.tables[table_name].columns.keys())
            keep = model_columns | derived_columns.get(table_name, set())
            # 같은 세션으로 컬럼을 조회해야 한다: inspect()는 별도 커넥션을 써서, 이 트랜잭션이
            # 방금 건 ALTER 락에 스스로 막히는 교착이 생긴다 (같은 커넥션은 자기 변경을 그냥 본다).
            live_columns = session.execute(
                text(f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}'")
            ).scalars().all()
            for column_name in live_columns:
                if column_name not in keep:
                    session.execute(text(f'ALTER TABLE {table_name} DROP COLUMN "{column_name}"'))
            legacy_uniques = session.execute(
                # 테이블명은 위 고정 목록에서만 오므로 f-string이 안전하다 (:param + ::cast는 text()가 오파싱).
                text(f"SELECT conname FROM pg_constraint WHERE contype = 'u' AND conrelid = '{table_name}'::regclass")
            ).all()
            for (constraint_name,) in legacy_uniques:
                session.execute(text(f'ALTER TABLE {table_name} DROP CONSTRAINT "{constraint_name}"'))
    session.commit()


def init_db() -> None:
    """스키마 생성 → 기준정보 시드 → 계약 검증. 두 백엔드가 같은 경로를 탄다."""
    DATA_DIR.mkdir(exist_ok=True)
    UPLOAD_DIR.mkdir(exist_ok=True)
    db_engine = configure_engine()
    Base.metadata.create_all(db_engine)  # ORM 모델이 스키마의 단일 출처
    with get_session() as session:
        resolve_postgres_type_drift(session)
        migrate_legacy_columns(session)
        # FK 순서: 표준계정(부모) → 체크리스트·별칭·양식(자식) → 문단 → 임베딩.
        ensure_reference_accounts(session)
        ensure_checklist_items(session)
        ensure_account_aliases(session)
        ensure_statement_templates(session)
        ensure_standards_paragraphs(session)
        ensure_paragraph_embeddings(session)
        ensure_vector_search(session)
        drop_derived_reference_tables(session)
        ensure_admin_user(session)
        session.commit()
        refresh_reference_cache(session)  # 기준정보 캐시 채우기 + 계약 검증(실패 시 시작 중단)


# ───────────────────────────────────────────────────────────────────────────
# §5 시드·기준정보 로더 (seeds/*.sql → ReferenceData)
# ───────────────────────────────────────────────────────────────────────────
PARAGRAPH_PUBLIC_COLUMNS = (
    StandardsParagraph.id,
    StandardsParagraph.standard_set,
    StandardsParagraph.reference_code,
    StandardsParagraph.paragraph_label,
    StandardsParagraph.account_key,
    StandardsParagraph.title,
    StandardsParagraph.content,
    StandardsParagraph.keywords,
)


def ensure_standards_paragraphs(session: Session) -> None:
    """기준서 문단을 seeds/standards_paragraphs.sql(단일 출처)에서 upsert한다.

    upsert(ON CONFLICT DO UPDATE)로 멱등성을 지키면서 embedding을 보존하고, content_hash가
    바뀐 행만 embedding을 비워 재임베딩 대상으로 표시한다. SQL 시드가 content와 content_hash를
    함께 제공하므로(gen_seeds로 일관 생성), 내용이 바뀌면 hash도 바뀌어 재임베딩이 트리거된다.
    """
    # 시드 실행 전 기존 해시를 기록해두고, 시드 후 해시가 바뀐 문단만 embedding을 비운다.
    before = dict(session.execute(select(StandardsParagraph.id, StandardsParagraph.content_hash)).all())
    # 시드 SQL은 embedding 컬럼을 건드리지 않으므로 기존 임베딩은 보존된다.
    run_seed(session, "standards_paragraphs")
    for paragraph_id, content_hash in session.execute(
        select(StandardsParagraph.id, StandardsParagraph.content_hash)
    ).all():
        if before.get(paragraph_id) != content_hash:
            session.execute(
                update(StandardsParagraph).where(StandardsParagraph.id == paragraph_id).values(embedding=None)
            )
    session.commit()


def find_standards_paragraphs(
    session: Session, account_key: str | None = None, query: str | None = None, standard_set: str | None = None
) -> list[dict]:
    stmt = select(*PARAGRAPH_PUBLIC_COLUMNS)
    if account_key:
        stmt = stmt.where(StandardsParagraph.account_key == account_key)
    if standard_set:
        stmt = stmt.where(StandardsParagraph.standard_set == standard_set)
    if query:
        like = f"%{query.strip()}%"
        stmt = stmt.where(
            StandardsParagraph.title.like(like)
            | StandardsParagraph.content.like(like)
            | StandardsParagraph.keywords.like(like)
            | StandardsParagraph.reference_code.like(like)
        )
    stmt = stmt.order_by(
        StandardsParagraph.account_key,
        StandardsParagraph.standard_set,
        StandardsParagraph.reference_code,
        StandardsParagraph.paragraph_label,
    )
    return [dict(row._mapping) for row in session.execute(stmt).all()]


def load_standards_paragraph_map(session: Session) -> dict:
    grouped: dict[str, list[dict]] = {}
    for row in find_standards_paragraphs(session):
        grouped.setdefault(row["account_key"], []).append(row)
    return grouped


def ensure_admin_user(session: Session) -> None:
    """ADMIN_EMAIL/ADMIN_PASSWORD가 설정되어 있으면 관리자 계정을 시드·갱신한다."""
    config = admin_config()
    if not config["configured"]:
        return
    password_hash = hash_password(os.environ.get("ADMIN_PASSWORD") or os.environ.get("GTF_ADMIN_PASSWORD") or "")
    read_only = bool(config["read_only"])

    admin = session.get(AppUser, ADMIN_USER_ID)
    if admin:
        # 같은 이메일을 쓰는 다른 계정이 있으면 UNIQUE 제약에 걸리므로 먼저 정리한다.
        conflicting = session.scalar(
            select(AppUser).where(AppUser.email == config["email"], AppUser.id != ADMIN_USER_ID)
        )
        if conflicting:
            session.execute(delete(UserSession).where(UserSession.user_id == conflicting.id))
            session.delete(conflicting)
            session.flush()
        admin.email = config["email"]
        admin.password_hash = password_hash
        admin.is_read_only = read_only
        session.commit()
        return

    existing = session.scalar(select(AppUser).where(AppUser.email == config["email"]))
    if existing:
        session.execute(delete(UserSession).where(UserSession.user_id == existing.id))
        session.delete(existing)
        session.flush()
    session.add(
        AppUser(
            id=ADMIN_USER_ID,
            email=config["email"],
            password_hash=password_hash,
            is_read_only=read_only,
            created_at=utc_now(),
        )
    )
    session.commit()


def demo_config() -> dict:
    enabled_env = os.environ.get("DEMO_LOGIN_ENABLED") or os.environ.get("GTF_DEMO_LOGIN_ENABLED") or "true"
    enabled = enabled_env.strip().lower() not in {"0", "false", "no", "off"}
    email = normalize_email(os.environ.get("DEMO_EMAIL") or os.environ.get("GTF_DEMO_EMAIL") or DEFAULT_DEMO_EMAIL)
    password = os.environ.get("DEMO_PASSWORD") or os.environ.get("GTF_DEMO_PASSWORD") or DEFAULT_DEMO_PASSWORD
    return {"enabled": enabled, "email": email, "password": password}


def ensure_demo_user(session: Session) -> dict | None:
    """데모(읽기 전용) 계정을 시드하고 공개용 dict로 돌려준다."""
    config = demo_config()
    if not config["enabled"] or not config["email"]:
        return None
    password_hash = hash_password(config["password"])
    user = session.scalar(select(AppUser).where(AppUser.email == config["email"]))
    if user:
        user.password_hash = password_hash
        user.is_read_only = True
    else:
        user = AppUser(
            id=DEMO_USER_ID,
            email=config["email"],
            password_hash=password_hash,
            is_read_only=True,
            created_at=utc_now(),
        )
        session.add(user)
    session.commit()
    return row_to_dict(user)


def run_seed(session: Session, name: str) -> None:
    """seeds/<name>.sql(SQLite/Postgres 공통 upsert 단일 문장)을 실행한다."""
    session.execute(text((SEED_DIR / f"{name}.sql").read_text(encoding="utf-8")))


def ensure_reference_accounts(session: Session) -> None:
    """표준계정을 seeds/standard_accounts.sql(단일 출처)에서 upsert한다.

    표준양식 라인·체크리스트·별칭 시드가 account_key 외래키를 참조하므로 먼저 실행되어야 한다.
    upsert라 재배포만으로 계정 카탈로그 확장이 라이브 DB에 반영된다.
    """
    run_seed(session, "standard_accounts")
    session.commit()


def ensure_checklist_items(session: Session) -> None:
    """판단 체크리스트 항목을 seeds/checklist_items.sql(단일 출처)에서 시드한다.

    다른 테이블이 참조하지 않으므로 전체 삭제 후 재삽입으로 코드와 일치시킨다.
    """
    session.execute(delete(ChecklistItem))
    run_seed(session, "checklist_items")
    session.commit()


def ensure_statement_templates(session: Session) -> None:
    """표준 재무제표 양식 라인 기준 테이블을 SQL 시드 파일(단일 출처)에서 upsert한다."""
    run_seed(session, "financial_statement_templates")
    session.commit()


def ensure_account_aliases(session: Session) -> None:
    """계정명 별칭 사전(kgaap_accounts)을 seeds/kgaap_accounts.sql(단일 출처)에서 시드한다.

    매칭 우선순위(긴 별칭 먼저)는 런타임에 별칭 길이로 계산하므로 DB에는 저장하지 않는다.
    표준계정 FK를 참조하므로 ensure_reference_accounts 뒤에 호출해야 한다. 옛 형식 id를
    남기지 않도록 전체 삭제 후 재삽입한다(참조하는 테이블 없음).
    """
    session.execute(delete(KgaapAccount))
    run_seed(session, "kgaap_accounts")
    session.commit()


def load_account_alias_map(session: Session) -> dict:
    """kgaap_accounts에서 별칭 → 계정키 맵을 로드한다 (매칭 순서는 normalize가 길이로 정함)."""
    rows = session.execute(
        select(KgaapAccount.kgaap_name, KgaapAccount.account_key)
        .where(KgaapAccount.active.is_(True))
        .order_by(KgaapAccount.kgaap_name)
    ).all()
    return {name: account_key for name, account_key in rows}


def load_statement_template_map(session: Session) -> dict:
    """계정키 → IFRS 표준양식 라인. generate_conversion이 표시 재무제표·라인명을 여기서 얻는다."""
    rows = session.execute(
        select(
            FinancialStatementTemplate.account_key,
            FinancialStatementTemplate.statement_type,
            FinancialStatementTemplate.section,
            FinancialStatementTemplate.line_item,
            FinancialStatementTemplate.display_order,
            FinancialStatementTemplate.basis,
        )
        .where(FinancialStatementTemplate.standard_set == "IFRS", FinancialStatementTemplate.active.is_(True))
        .order_by(FinancialStatementTemplate.display_order)
    ).all()
    return {row.account_key: dict(row._mapping) for row in rows}


def load_reference_data(session: Session) -> ReferenceData:
    """DB 기준정보 테이블을 읽어 domain에 주입할 ReferenceData로 묶는다 (코드에 하드코딩 없음)."""
    accounts = {
        account.account_key: {
            "code": account.standard_code,
            "label": account.internal_label,
            "ifrs": account.ifrs_account,
            "type": account.mapping_type,
            "rule": account.rule_summary,
        }
        for account in session.scalars(select(StandardAccount))
    }
    checklists: dict[str, list] = {}
    for item in session.scalars(
        select(ChecklistItem).order_by(ChecklistItem.account_key, ChecklistItem.display_order)
    ):
        checklists.setdefault(item.account_key, []).append(
            {
                "key": item.item_key,
                "label": item.label,
                "type": item.input_type,
                "options": [option.strip() for option in (item.options or "").split("/") if option.strip()],
                "required": bool(item.required),
            }
        )
    return ReferenceData(
        accounts=accounts,
        aliases=load_account_alias_map(session),
        checklists=checklists,
        templates=load_statement_template_map(session),
        paragraphs=load_standards_paragraph_map(session),
    )


def sort_statements_by_code(statements: list[dict]) -> list[dict]:
    """계정 행을 계정코드 기반 표시 순서(자산 → 부채 → 자본 → 손익)로 정렬한다.

    SQL의 알파벳 정렬은 자본(E)이 부채(L)보다 앞서 어긋나므로 코드에서 도출한 순서로 정렬한다.
    변환 결과(entries)와 동일한 기준이라 매핑 테이블·검토 화면·리포트가 같은 순서로 보인다.
    """
    return sorted(
        statements,
        key=lambda s: (account_presentation_order(str(s.get("standard_code") or "")), str(s.get("created_at") or "")),
    )


def drop_derived_reference_tables(session: Session) -> None:
    """예전 배포가 만든 파생 조회 테이블(ifrs_accounts/mapping_rules/standards_references)을 정리한다.

    이 테이블들은 standard_accounts·checklist_items·standards_paragraphs를 부팅 때마다 복제한
    그림자였고 변환 로직은 읽지 않았다. 원본에서 직접 조회하도록 정리했으므로, 라이브 DB에
    남은 잔재만 지운다(create_all은 테이블을 삭제하지 않는다). 새 DB에서는 no-op이다.
    """
    for orphan in ("ifrs_accounts", "mapping_rules", "standards_references"):
        session.execute(text(f"DROP TABLE IF EXISTS {orphan}"))
    session.commit()


# ───────────────────────────────────────────────────────────────────────────
# §6 감사로그·행 직렬화
# ───────────────────────────────────────────────────────────────────────────
def row_to_dict(obj) -> dict:
    """ORM 인스턴스나 Row를 평범한 dict로 바꾼다 (JSON 응답용)."""
    if obj is None:
        return {}
    if hasattr(obj, "_mapping"):  # SQLAlchemy Row
        return {key: normalize_db_value(value) for key, value in obj._mapping.items()}
    mapper = inspect(type(obj))
    return {column.key: normalize_db_value(getattr(obj, column.key)) for column in mapper.column_attrs}


def database_ready() -> bool:
    try:
        with get_session() as session:
            session.execute(text("SELECT 1"))
        return True
    except Exception:
        return False


def load_project_statements(session: Session, project_id: str) -> list[dict]:
    """프로젝트의 계정행을 표시 순서로, 체크리스트 JSON을 파싱해 로드한다 (조회·변환·내보내기 공용)."""
    return sort_statements_by_code([
        dict(row_to_dict(statement), checklist=parse_json_field(statement.checklist_json, []))
        for statement in session.scalars(select(Statement).where(Statement.project_id == project_id))
    ])


def ai_classification_audit(ai_classification: dict) -> dict:
    """AI 1차 분류 결과를 감사로그용 요약으로 축약한다 (추출 3개 경로 공용)."""
    return {
        "status": ai_classification.get("status"),
        "model": ai_classification.get("model"),
        "suggested_accounts": sorted((ai_classification.get("suggestions") or {}).keys()),
        "human_review_required": True,
    }


def log_event(session: Session, project_id: str, event_type: str, detail: dict, actor: str = "system") -> None:
    session.add(
        AuditLog(
            id=str(uuid.uuid4()),
            project_id=project_id,
            event_type=event_type,
            actor=actor,
            detail_json=json.dumps(detail, ensure_ascii=False),
            created_at=utc_now(),
        )
    )


# ───────────────────────────────────────────────────────────────────────────
# §7 파일 파서 (xlsx / pdf)
# ───────────────────────────────────────────────────────────────────────────
def read_xlsx_table(path: Path) -> list[list[str]]:
    """첫 워크시트를 문자열 표로 읽는다. 쓰기(excel_export)와 같은 openpyxl을 사용한다."""
    workbook = load_workbook(path, read_only=True, data_only=True)  # data_only: 수식 대신 계산값
    try:
        sheet = workbook.worksheets[0]
        return [
            ["" if cell is None else str(cell).strip() for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()


def parse_xlsx_statement_rows(path: Path) -> tuple[list[dict], list[str]]:
    table = read_xlsx_table(path)
    rows = []
    issues = []
    header = None
    account_index = None
    amount_index = None

    for row_index, row in enumerate(table):
        compact = [str(cell).strip() for cell in row]
        lowered = [cell.replace(" ", "").lower() for cell in compact]
        account_candidates = [i for i, cell in enumerate(lowered) if cell in {"계정명", "계정", "account", "accountname"}]
        amount_candidates = [i for i, cell in enumerate(lowered) if cell in {"금액", "amount", "당기", "current"}]
        if account_candidates and amount_candidates:
            header = row_index
            account_index = account_candidates[0]
            amount_index = amount_candidates[-1]
            break

    data_rows = table[header + 1 :] if header is not None else table
    for row in data_rows:
        compact = [str(cell).strip() for cell in row]
        if not any(compact):
            continue
        if account_index is not None and amount_index is not None and len(compact) > max(account_index, amount_index):
            name = compact[account_index]
            amount_value = compact[amount_index]
        else:
            text_cells = [cell for cell in compact if cell and not looks_numeric(cell)]
            numeric_cells = [cell for cell in compact if looks_numeric(cell)]
            if not text_cells or not numeric_cells:
                continue
            name = text_cells[0]
            amount_value = numeric_cells[-1]
        if not name or "계정" in name.replace(" ", "") or "금액" == name.replace(" ", ""):
            continue
        rows.append({"account_name": name, "amount": parse_amount(amount_value)})

    if not rows:
        issues.append("Excel 파일은 읽었지만 계정명/금액 행을 찾지 못했습니다.")
    return rows, issues


def clean_pdf_account_name(value) -> str:
    text = re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()
    text = re.sub(r"\(주석\s*[^)]*\)", "", text).strip()
    text = re.sub(r"^[ㆍ·]\s*", "", text)
    text = re.sub(r"^(?:[0-9]+\.|\([0-9]+\)|[가-힣]\.|[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩXI]+\.?)\s*", "", text).strip()
    return text


def pdf_statement_account_key(name: str) -> str:
    compact = re.sub(r"\s+", "", name)
    if not compact:
        return "other"
    exclusions = [
        "매출원가",
        "매출총이익",
        "영업활동",
        "현금흐름",
        "현금의유입",
        "현금의유출",
        "현금유입",
        "현금유출",
        "활동으로인한",
        "감소",
        "증가",
        "기초",
        "기말",
        "감가상각",
        "상각비",
        "처분손실",
        "처분이익",
        "당기순",
        "주당",
        "전환사채발행",
        "유상증자",
        "주식선택권",
        "토지재평가",
        "영업외수익",
        "이자수익",
    ]
    if any(token in compact for token in exclusions):
        return "other"
    return normalize_account_name(name, REFERENCE.aliases)


def pdf_row_current_amount(row: list) -> float | None:
    for cell in row[1:]:
        if looks_numeric(cell):
            return parse_amount(cell)
    return None


def parse_pdf_text_statement_rows(text: str) -> list[dict]:
    rows: list[dict] = []
    account_keywords = {
        "cash": ["현금및현금성자산", "현금및 현금성자산", "현금성자산"],
        "receivables": ["매출채권", "미수금", "미수수익", "계약자산"],
        "inventory": ["재고자산", "상품", "제품", "원재료"],
        "lease": ["사용권자산", "리스부채"],
        "development": ["개발비", "무형자산"],
        "revenue": ["매출액", "매출", "영업수익", "수익"],
        "financial_instrument": ["금융자산", "금융부채", "단기차입금", "장기차입금", "전환사채", "파생상품"],
        "provision": ["충당부채", "판매보증충당부채", "복구충당부채"],
    }
    amount_pattern = r"\(?-?\d{1,3}(?:,\d{3})+(?:\.\d+)?\)?|\(?-?\d{5,}(?:\.\d+)?\)?"
    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        if not line or len(line) > 180:
            continue
        compact = re.sub(r"\s+", "", line)
        matched_name = ""
        for keywords in account_keywords.values():
            matched_name = next((keyword for keyword in keywords if keyword.replace(" ", "") in compact), "")
            if matched_name:
                break
        if not matched_name:
            continue
        amounts = re.findall(amount_pattern, line)
        if not amounts:
            continue
        rows.append({"account_name": matched_name, "amount": parse_amount(amounts[0])})
    return rows


def parse_pdf_statement_rows(path: Path) -> tuple[list[dict], list[str]]:
    try:
        import pdfplumber
    except ImportError:
        return [], ["PDF 표 직접 추출을 위해 pdfplumber 패키지가 필요합니다."]

    rows: list[dict] = []
    issues: list[str] = []
    started = False
    pages_scanned = 0

    with pdfplumber.open(path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            compact_text = re.sub(r"\s+", "", text)
            if started and "재무제표주석" in compact_text:
                break
            if any(marker in compact_text for marker in ("재무상태표", "손익계산서", "현금흐름표", "자본변동표")):
                started = True
            if not started:
                continue

            pages_scanned += 1
            for row in parse_pdf_text_statement_rows(text):
                rows.append(row)
            for table in page.extract_tables() or []:
                for raw_row in table:
                    if not raw_row:
                        continue
                    name = clean_pdf_account_name(raw_row[0])
                    if not name or name in {"과목", "과 목", "구분", "자산", "자 산", "부채", "자본"}:
                        continue
                    account_key = pdf_statement_account_key(name)
                    if account_key == "other":
                        continue
                    amount = pdf_row_current_amount(raw_row)
                    if amount is None:
                        continue
                    rows.append({"account_name": name, "amount": amount})

    deduped: list[dict] = []
    seen: set[tuple[str, float]] = set()
    for row in rows:
        key = (row["account_name"], float(row["amount"]))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)

    if pages_scanned:
        issues.append(f"PDF 표 직접 추출: 재무제표 본문 {pages_scanned}쪽을 분석했습니다.")
    if not deduped:
        issues.append("PDF 표는 읽었지만 변환 가능한 핵심 계정 행을 찾지 못했습니다.")
    return deduped, issues


# ───────────────────────────────────────────────────────────────────────────
# §8 외부 서비스 설정 (OCR·AI·DART)
# ───────────────────────────────────────────────────────────────────────────
def ocr_config() -> dict:
    provider = os.environ.get("OCR_PROVIDER", "gemini").strip() or "gemini"
    model = os.environ.get("GEMINI_OCR_MODEL", "gemini-3.5-flash").strip() or "gemini-3.5-flash"
    api_key_source = "environment" if os.environ.get("GEMINI_API_KEY") else "none"
    api_key_ready = api_key_source != "none"
    return {
        "provider": provider,
        "model": model,
        "api_key_ready": api_key_ready,
        "api_key_source": api_key_source,
        "mode": "connected" if api_key_ready else "placeholder",
        "manual_review_on_failure": True,
    }


def ai_config() -> dict:
    model = os.environ.get("OPENAI_MODEL", OPENAI_DEFAULT_MODEL).strip() or OPENAI_DEFAULT_MODEL
    api_key_source = "environment" if os.environ.get("OPENAI_API_KEY") else "none"
    api_key_ready = api_key_source != "none"
    return {
        "provider": "openai",
        "model": model,
        "api_key_ready": api_key_ready,
        "api_key_source": api_key_source,
        "mode": "connected" if api_key_ready else "not_configured",
        "human_review_required": True,
    }


def dart_config() -> dict:
    return {
        "provider": "opendart",
        "api_key_ready": bool(os.environ.get("DART_API_KEY", "").strip()),
        "corp_code_lookup": True,
        "supported_reports": {
            "11013": "1분기보고서",
            "11012": "반기보고서",
            "11014": "3분기보고서",
            "11011": "사업보고서",
        },
        "supported_fs_div": {"CFS": "연결", "OFS": "별도"},
    }


def supported_ocr_mime(path: Path, content_type: str) -> str | None:
    suffix = path.suffix.lower()
    if suffix == ".pdf" or "pdf" in content_type:
        return "application/pdf"
    if suffix in {".png", ".jpg", ".jpeg"}:
        return "image/png" if suffix == ".png" else "image/jpeg"
    guessed, _ = mimetypes.guess_type(path.name)
    if guessed in {"application/pdf", "image/png", "image/jpeg"}:
        return guessed
    return None


def extract_json_object(text: str) -> dict:
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", cleaned, re.S)
        if not match:
            raise
        return json.loads(match.group(0))


def gemini_response_text(response: dict) -> str:
    if isinstance(response.get("output_text"), str):
        return response["output_text"]
    if isinstance(response.get("text"), str):
        return response["text"]
    texts: list[str] = []
    json_like_texts: list[str] = []

    def collect_text(value) -> None:
        if isinstance(value, dict):
            for key in ("output_text", "text"):
                if isinstance(value.get(key), str) and value[key].strip():
                    texts.append(value[key])
            for item in value.values():
                if isinstance(item, str) and '"rows"' in item and "account_name" in item:
                    json_like_texts.append(item)
            for key in ("model_output", "content", "parts", "steps", "output", "outputs"):
                collect_text(value.get(key))
        elif isinstance(value, list):
            for item in value:
                collect_text(item)

    collect_text(response.get("model_output"))
    collect_text(response.get("steps"))
    collect_text(response.get("output"))
    if texts:
        return "\n".join(texts).strip()
    if json_like_texts:
        return "\n".join(json_like_texts).strip()

    candidates = response.get("candidates") or []
    if not candidates:
        return ""
    parts = candidates[0].get("content", {}).get("parts") or []
    candidate_texts = [part.get("text", "") for part in parts if isinstance(part.get("text"), str)]
    return "\n".join(candidate_texts).strip()


# ───────────────────────────────────────────────────────────────────────────
# §9 AI·RAG
# ───────────────────────────────────────────────────────────────────────────
def openai_embed(texts: list[str]) -> list[list[float]] | None:
    """텍스트 목록을 OpenAI 임베딩 벡터로 변환한다.

    OPENAI_API_KEY가 없거나 호출이 실패하면 None을 반환하고, 검색은 키워드 방식으로 폴백한다.
    문단 수가 적어 한 번의 배치 호출로 처리한다.
    """
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    texts = [str(t) for t in (texts or []) if str(t).strip()]
    if not api_key or not texts:
        return None
    try:
        client = OpenAI(api_key=api_key, timeout=45)
        result = client.embeddings.create(model=EMBEDDING_MODEL, input=texts)
        vectors = [item.embedding for item in result.data]
        return vectors if len(vectors) == len(texts) else None
    except openai.OpenAIError:
        return None


def _cosine_similarity(a: list[float], b: list[float]) -> float:
    dot = sum(x * y for x, y in zip(a, b))
    norm_a = math.sqrt(sum(x * x for x in a))
    norm_b = math.sqrt(sum(y * y for y in b))
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return dot / (norm_a * norm_b)


def ensure_paragraph_embeddings(session: Session) -> None:
    """기준서 문단을 OpenAI 임베딩으로 변환해 embedding 컬럼(이식형 JSON)에 저장한다 (RAG 검색용).

    이 텍스트 컬럼이 벡터의 단일 출처다. Postgres에서는 ensure_vector_search가 이를 vector 타입
    파생 컬럼(embedding_vec)으로 캐스팅해 DB 안에서 KNN 검색하고, SQLite에서는 앱의 코사인으로 검색한다.

    embedding이 비어 있는 행(신규 또는 내용이 바뀐 문단)만 임베딩한다. 내용이 그대로면
    OpenAI 호출을 아예 하지 않으므로 서버 재시작 비용이 0이다.
    OPENAI_API_KEY가 없거나 호출이 실패하면 임베딩을 비워 두고 키워드 검색으로 폴백하며,
    서버 시작을 막지 않도록 실패를 조용히 무시한다.
    """
    rows = session.execute(
        select(
            StandardsParagraph.id,
            StandardsParagraph.reference_code,
            StandardsParagraph.title,
            StandardsParagraph.content,
        )
        .where(StandardsParagraph.embedding.is_(None))
        .order_by(StandardsParagraph.id)
    ).all()
    if not rows:
        return  # 재임베딩할 문단이 없음 → OpenAI 호출 0회
    texts = [f"{reference_code} {title} {content}" for _id, reference_code, title, content in rows]
    vectors = openai_embed(texts)
    if not vectors:
        return
    for (paragraph_id, *_), vector in zip(rows, vectors):
        session.execute(
            update(StandardsParagraph)
            .where(StandardsParagraph.id == paragraph_id)
            .values(embedding=json.dumps(vector))
        )
    session.commit()


# Postgres에서 pgvector 가속이 준비됐는지 (부팅 시 ensure_vector_search가 설정).
# 벡터의 진실 원천은 embedding TEXT(JSON) 컬럼이고, embedding_vec은 파생 인덱스라
# 언제든 재구축 가능하다. 확장이 없는 Postgres에서는 파이썬 코사인으로 폴백한다.
PGVECTOR_READY = False


def ensure_vector_search(session: Session) -> None:
    """Postgres면 pgvector 확장·파생 벡터 컬럼·HNSW 인덱스를 준비한다.

    embedding(TEXT JSON)이 진실 원천이고 embedding_vec은 그것을 ::vector로 캐스팅한
    파생 컬럼이다. 문단 임베딩이 갱신된 뒤 호출되어 파생 컬럼을 동기화한다.
    확장 설치가 불가능한 환경이면 조용히 건너뛰고 파이썬 코사인 검색을 그대로 쓴다.
    """
    global PGVECTOR_READY
    PGVECTOR_READY = False
    if backend() != "postgres":
        return
    sample = session.scalar(
        select(StandardsParagraph.embedding).where(StandardsParagraph.embedding.is_not(None)).limit(1)
    )
    if not sample:
        return  # 임베딩이 아직 없으면(키 미설정 등) 다음 부팅에서 준비한다
    dims = len(json.loads(sample))
    try:
        session.execute(text("CREATE EXTENSION IF NOT EXISTS vector"))
        session.execute(text(f"ALTER TABLE standards_paragraphs ADD COLUMN IF NOT EXISTS embedding_vec vector({dims})"))
        # 파생 컬럼을 원본(embedding TEXT)에서 무조건 재캐스팅한다 — 부팅당 1회, 수만 행까지도 수 초.
        # 조건부 동기화의 staleness 버그 가능성보다 무조건 갱신의 단순함을 택했다.
        session.execute(text(
            "UPDATE standards_paragraphs SET embedding_vec = embedding::vector WHERE embedding IS NOT NULL"
        ))
        # 의도적으로 ANN 인덱스(HNSW)를 만들지 않는다: 인덱스 없는 <=> 정렬은 "정확한" 전수 KNN이며
        # 이 규모(수백 문단)에선 ms 단위다. HNSW는 근사라 재현율을 희생하므로 수만 문단부터 추가한다.
        session.commit()
        PGVECTOR_READY = True
    except Exception:
        session.rollback()  # 확장 미지원 등 — 파이썬 코사인 폴백으로 동작


def pgvector_search(session: Session, query_vector: list[float], account_key: str | None, standard_set: str | None, k: int) -> list[dict]:
    """pgvector KNN: 코사인 거리 연산자(<=>)로 DB 안에서 top-k를 뽑는다."""
    conditions = ["embedding_vec IS NOT NULL"]
    params: dict = {"qv": json.dumps(query_vector), "k": k}
    if account_key:
        conditions.append("account_key = :ak")
        params["ak"] = account_key
    if standard_set:
        conditions.append("standard_set = :ss")
        params["ss"] = standard_set
    rows = session.execute(text(
        "SELECT id, standard_set, reference_code, paragraph_label, account_key, title, content, keywords, "
        "1 - (embedding_vec <=> CAST(:qv AS vector)) AS similarity "
        f"FROM standards_paragraphs WHERE {' AND '.join(conditions)} "
        "ORDER BY embedding_vec <=> CAST(:qv AS vector) LIMIT :k"
    ), params).all()
    results = []
    for row in rows:
        para = dict(row._mapping)
        para["similarity"] = round(float(para["similarity"]), 4)
        para["retrieval"] = "semantic"
        results.append(para)
    return results


def semantic_search_paragraphs(
    session: Session, query: str, account_key: str | None = None, standard_set: str | None = None, k: int = 5
) -> list[dict]:
    """질의를 임베딩해 코사인 유사도로 기준서 문단을 검색한다 (RAG 검색 단계).

    임베딩이 없으면(키 미설정·미생성) 키워드 LIKE 검색으로 폴백한다.
    반환 문단에는 retrieval 방식(semantic/keyword)과 유사도가 표시된다.
    """
    def keyword_fallback() -> list[dict]:
        # 임베딩이 없을 때의 폴백: 전체 문구가 안 맞으면 토큰 단위로도 매칭한다.
        seen: dict[str, dict] = {}
        for token in [query, *query.split()]:
            token = (token or "").strip()
            if not token:
                continue
            for row in find_standards_paragraphs(session, account_key=account_key, query=token, standard_set=standard_set):
                seen.setdefault(row["id"], dict(row, retrieval="keyword"))
            if len(seen) >= k:
                break
        if not seen and not query:
            for row in find_standards_paragraphs(session, account_key=account_key, standard_set=standard_set):
                seen.setdefault(row["id"], dict(row, retrieval="keyword"))
        return list(seen.values())[:k]

    query_vectors = openai_embed([query]) if query else None
    if not query_vectors:
        return keyword_fallback()
    query_vector = query_vectors[0]

    if PGVECTOR_READY:
        results = pgvector_search(session, query_vector, account_key, standard_set, k)
        if results:
            return results
        return keyword_fallback()

    stmt = select(*PARAGRAPH_PUBLIC_COLUMNS, StandardsParagraph.embedding)
    if account_key:
        stmt = stmt.where(StandardsParagraph.account_key == account_key)
    if standard_set:
        stmt = stmt.where(StandardsParagraph.standard_set == standard_set)

    scored: list[tuple[float, dict]] = []
    for row in session.execute(stmt).all():
        para = dict(row._mapping)
        embedding = para.pop("embedding", None)
        if not embedding:
            continue
        try:
            vector = json.loads(embedding)
        except (ValueError, TypeError):
            continue
        scored.append((_cosine_similarity(query_vector, vector), para))

    if not scored:
        return keyword_fallback()
    scored.sort(key=lambda pair: pair[0], reverse=True)
    results = []
    for score, para in scored[:k]:
        para["similarity"] = round(score, 4)
        para["retrieval"] = "semantic"
        results.append(para)
    return results


def call_ai_judgment(project: dict, entries: list[dict], judgment_items: list[dict], retrieved_context: list[dict] | None = None) -> dict:
    config = ai_config()
    if not judgment_items:
        return {
            "provider": "openai",
            "model": config["model"],
            "status": "skipped",
            "items": [],
            "overall_note": "판단 필요 항목이 없어 OpenAI 판단 보조를 건너뛰었습니다.",
            "human_review_required": True,
        }
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return {
            "provider": "openai",
            "model": config["model"],
            "status": "not_configured",
            "items": [],
            "overall_note": "OPENAI_API_KEY가 서버 환경변수에 설정되지 않아 규정 근거 요약은 생성하지 않았습니다.",
            "issues": ["OPENAI_API_KEY가 서버 환경변수에 설정되지 않았습니다."],
            "human_review_required": True,
        }

    compact_entries = [
        {
            "source_account": entry.get("source_account"),
            "standard_code": entry.get("standard_code"),
            "target_account": entry.get("target_account"),
            "statement_line_item": entry.get("statement_line_item"),
            "mapping_type": entry.get("mapping_type"),
            "basis": entry.get("basis"),
            "calculation": entry.get("calculation"),
        }
        for entry in entries
        if entry.get("mapping_type") == "judgment"
    ]
    prompt = {
        "project": {
            "company_name": project.get("company_name"),
            "period": project.get("period"),
            "source_standard": project.get("source_standard"),
            "target_standard": project.get("target_standard"),
        },
        "judgment_entries": compact_entries,
        "checklist_inputs": judgment_items,
        "retrieved_standards": retrieved_context or [],
        "response_contract": {
            "items": [
                {
                    "account": "계정명",
                    "risk_level": "low|medium|high",
                    "classification_hint": "검토자가 확인할 분류 방향",
                    "additional_questions": ["추가로 확인할 질문"],
                    "review_note": "사람 검토자가 볼 짧은 검토 메모",
                    "basis_summary": "①판단 쟁점 ②관련 기준서 문단의 요지(reference_code 인용) ③검토 결론 방향, 2~3문장",
                }
            ],
            "overall_note": "전체 검토 메모",
        },
    }
    payload = {
        "model": config["model"],
        # 판단항목이 10개 이상일 수 있어 항목당 근거가 잘리지 않도록 넉넉히 잡는다.
        "max_output_tokens": 3000,
        "instructions": (
            "너는 K-GAAP 재무제표를 IFRS 초안으로 변환하는 회계 검토 보조자다. "
            "최종 회계처리를 확정하지 말고, 사용자가 입력한 체크리스트와 변환 초안을 바탕으로 "
            "판단 필요 항목, 추가 질문, 기준 근거 요약만 한국어로 제시한다. "
            "basis_summary는 한국어 2~3문장으로, ①이 계정의 판단 쟁점 ②관련 기준서 문단의 요지 ③검토 결론 방향의 순서로 쓴다. "
            "반드시 retrieved_standards로 제공된 기준서 문단에 근거해 작성하고 "
            "해당 문단의 reference_code를 문장 안에 인용하며, 제공된 문단에 없는 내용은 추측하지 않는다. "
            "additional_questions에는 검토자가 실제로 확인해야 할 구체적 질문을 1개 이상 담는다. "
            "금액은 절대 새로 계산하지 않는다. "
            "반드시 사람이 최종 검토하고 승인해야 한다는 전제를 유지한다. JSON만 반환한다."
        ),
        "input": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": (
                            "다음 변환 초안의 판단 필요 항목을 검토해 JSON만 반환하세요. "
                            "마크다운과 설명 문장은 쓰지 마세요.\n"
                            + json.dumps(prompt, ensure_ascii=False)
                        ),
                    }
                ],
            }
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": "gtf_judgment_assistance",
                "schema": {
                    "type": "object",
                    "properties": {
                        "items": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "account": {"type": "string"},
                                    "risk_level": {"type": "string"},
                                    "classification_hint": {"type": "string"},
                                    "additional_questions": {"type": "array", "items": {"type": "string"}},
                                    "review_note": {"type": "string"},
                                    "basis_summary": {"type": "string"},
                                },
                                "required": [
                                    "account",
                                    "risk_level",
                                    "classification_hint",
                                    "additional_questions",
                                    "review_note",
                                    "basis_summary",
                                ],
                                "additionalProperties": False,
                            },
                        },
                        "overall_note": {"type": "string"},
                    },
                    "required": ["items", "overall_note"],
                    "additionalProperties": False,
                },
                "strict": True,
            }
        },
    }
    try:
        client = OpenAI(api_key=api_key, timeout=45)
        response = client.responses.create(**payload)
    except openai.APIStatusError as exc:
        return {
            "provider": "openai",
            "model": config["model"],
            "status": "failed",
            "items": [],
            "overall_note": "OpenAI 판단 보조 요청이 실패했습니다. 변환 초안은 저장되며 사람이 검토해야 합니다.",
            "issues": [f"OpenAI 요청 실패: HTTP {exc.status_code}", exc.response.text[:500]],
            "human_review_required": True,
        }
    except openai.APITimeoutError:
        return {
            "provider": "openai",
            "model": config["model"],
            "status": "failed",
            "items": [],
            "overall_note": "OpenAI 판단 보조 요청 시간이 초과되었습니다.",
            "issues": ["OpenAI 요청 시간이 초과되었습니다."],
            "human_review_required": True,
        }
    except openai.OpenAIError as exc:
        return {
            "provider": "openai",
            "model": config["model"],
            "status": "failed",
            "items": [],
            "overall_note": "OpenAI 판단 보조 네트워크 오류가 발생했습니다.",
            "issues": [f"OpenAI 네트워크 오류: {exc}"],
            "human_review_required": True,
        }

    text = (response.output_text or "").strip()
    if not text:
        return {
            "provider": "openai",
            "model": config["model"],
            "status": "failed",
            "items": [],
            "overall_note": "OpenAI 응답에서 검토 텍스트를 찾지 못했습니다.",
            "issues": ["OpenAI 응답 텍스트가 비어 있습니다."],
            "human_review_required": True,
        }
    try:
        parsed = extract_json_object(text)
    except json.JSONDecodeError:
        return {
            "provider": "openai",
            "model": config["model"],
            "status": "failed",
            "items": [],
            "overall_note": "OpenAI 응답을 JSON으로 해석하지 못했습니다.",
            "issues": ["OpenAI 응답 JSON 해석 실패", text[:500]],
            "human_review_required": True,
        }

    items = parsed.get("items") if isinstance(parsed.get("items"), list) else []
    retrieval_mode = "none"
    if retrieved_context:
        modes = {p.get("retrieval") for ctx in retrieved_context for p in ctx.get("paragraphs", [])}
        retrieval_mode = "semantic" if "semantic" in modes else ("keyword" if "keyword" in modes else "none")
    return {
        "provider": "openai",
        "model": config["model"],
        "status": "connected",
        "items": items,
        "overall_note": str(parsed.get("overall_note") or "사람 검토와 승인이 필요합니다."),
        "retrieved_context": retrieved_context or [],
        "retrieval_mode": retrieval_mode,
        "human_review_required": True,
    }


def call_ai_classification(unmapped_accounts: list[str], session: Session | None = None) -> dict:
    """키워드 매핑에 실패한 계정명에 대해 표준계정 후보를 제안받는 AI 1차 분류.

    제안은 추출 결과에 후보로만 저장되며, 담당자가 추출 결과를 반영하는 시점에
    확정된다. 분류를 자동 확정하지 않는다.

    판단보조(call_ai_judgment)와 같은 원칙으로 근거를 접지(grounding)한다:
    계정명마다 관련 기준서 문단을 검색해 프롬프트에 넣고, 근거는 제공된 문단의
    reference_code를 인용해서만 쓰게 한다. 재료 없이 지어내는 빈약한 한 줄
    근거와 환각 인용을 함께 막는다.
    """
    config = ai_config()
    base = {"provider": "openai", "model": config["model"], "human_review_required": True}
    if not unmapped_accounts:
        return {**base, "status": "skipped", "suggestions": {}, "note": "표준코드 매핑에 실패한 계정이 없어 AI 1차 분류를 건너뛰었습니다."}
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return {
            **base,
            "status": "not_configured",
            "suggestions": {},
            "note": "OPENAI_API_KEY가 설정되지 않아 미분류 계정은 담당자가 직접 분류해야 합니다.",
            "issues": ["OPENAI_API_KEY가 서버 환경변수에 설정되지 않았습니다."],
        }

    candidates = [
        {"account_key": key, "label": account["label"], "ifrs": account["ifrs"], "rule": account["rule"]}
        for key, account in REFERENCE.accounts.items()
        if key != "other"
    ]
    # 계정명마다 관련 기준서 문단을 검색해 근거 재료로 프롬프트에 첨부한다 (RAG).
    retrieved_standards = {}
    if session is not None:
        for name in unmapped_accounts:
            paragraphs = semantic_search_paragraphs(session, name, k=3)
            retrieved_standards[name] = [
                {
                    "reference_code": p.get("reference_code"),
                    "paragraph_label": p.get("paragraph_label"),
                    "title": p.get("title"),
                    "content": p.get("content"),
                }
                for p in paragraphs
            ]
    prompt = {
        "unmapped_accounts": unmapped_accounts,
        "candidate_standard_accounts": candidates,
        "retrieved_standards": retrieved_standards,
        "instruction": "각 계정명을 후보 표준계정 중 가장 적합한 하나로 분류 제안하세요. 확신이 없으면 suggested_account_key를 'other'로 두세요.",
    }
    payload = {
        "model": config["model"],
        "max_output_tokens": 2000,
        "instructions": (
            "너는 K-GAAP 재무제표 계정을 내부 표준계정으로 분류하는 회계 보조자다. "
            "규칙 기반 매핑에 실패한 계정명에 대해 후보 표준계정 중 하나를 제안한다. "
            "rationale은 한국어 2~3문장으로, ①계정의 경제적 성격 ②적용되는 기준 ③이 표준계정이 적합한 이유의 순서로 쓴다. "
            "basis_reference에는 retrieved_standards로 제공된 문단의 reference_code만 인용하고, "
            "관련 문단이 없으면 빈 문자열로 둔다. 제공되지 않은 기준서는 절대 지어내지 않는다. "
            "alternative_account_key에는 두 번째로 유력했던 후보를, alternative_rejected_reason에는 그 후보를 배제한 이유를 한 문장으로 쓴다. "
            "분류를 확정하지 말고 제안만 하며, 확신이 없으면 other를 반환한다. JSON만 반환한다."
        ),
        "input": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": "다음 미분류 계정을 분류 제안해 JSON만 반환하세요.\n" + json.dumps(prompt, ensure_ascii=False),
                    }
                ],
            }
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": "gtf_account_classification",
                "schema": {
                    "type": "object",
                    "properties": {
                        "items": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "account_name": {"type": "string"},
                                    "suggested_account_key": {"type": "string"},
                                    "confidence": {"type": "string"},
                                    "rationale": {"type": "string"},
                                    "basis_reference": {"type": "string"},
                                    "alternative_account_key": {"type": "string"},
                                    "alternative_rejected_reason": {"type": "string"},
                                },
                                "required": ["account_name", "suggested_account_key", "confidence", "rationale", "basis_reference", "alternative_account_key", "alternative_rejected_reason"],
                                "additionalProperties": False,
                            },
                        }
                    },
                    "required": ["items"],
                    "additionalProperties": False,
                },
                "strict": True,
            }
        },
    }
    try:
        client = OpenAI(api_key=api_key, timeout=45)
        response = client.responses.create(**payload)
    except openai.APIStatusError as exc:
        return {
            **base,
            "status": "failed",
            "suggestions": {},
            "note": "AI 1차 분류 요청이 실패해 미분류 계정은 담당자가 직접 분류해야 합니다.",
            "issues": [f"OpenAI 분류 요청 실패: HTTP {exc.status_code}", exc.response.text[:500]],
        }
    except openai.OpenAIError as exc:
        return {
            **base,
            "status": "failed",
            "suggestions": {},
            "note": "AI 1차 분류 네트워크 오류가 발생했습니다.",
            "issues": [f"OpenAI 분류 네트워크 오류: {exc}"],
        }

    text = (response.output_text or "").strip()
    try:
        parsed = extract_json_object(text) if text else {}
    except json.JSONDecodeError:
        parsed = {}
    items = parsed.get("items") if isinstance(parsed.get("items"), list) else []
    suggestions = {}
    for item in items:
        name = str(item.get("account_name") or "").strip()
        suggested_key = str(item.get("suggested_account_key") or "").strip()
        if not name or suggested_key not in REFERENCE.accounts or suggested_key == "other":
            continue
        alternative_key = str(item.get("alternative_account_key") or "").strip()
        suggestions[name] = {
            "account_key": suggested_key,
            "label": REFERENCE.accounts[suggested_key]["label"],
            "confidence": str(item.get("confidence") or "unknown"),
            "rationale": str(item.get("rationale") or "").strip(),
            "basis_reference": str(item.get("basis_reference") or "").strip(),
            "alternative_label": REFERENCE.accounts[alternative_key]["label"] if alternative_key in REFERENCE.accounts else "",
            "alternative_rejected_reason": str(item.get("alternative_rejected_reason") or "").strip(),
            "provider": "openai",
            "model": config["model"],
            "human_review_required": True,
        }
    return {
        **base,
        "status": "connected",
        "suggestions": suggestions,
        "note": f"미분류 계정 {len(unmapped_accounts)}건 중 {len(suggestions)}건에 AI 1차 분류 제안이 생성되었습니다. 반영 시 담당자 확정이 필요합니다.",
    }


def attach_ai_classification(rows: list[dict], session: Session | None = None) -> tuple[list[dict], dict]:
    """추출 행 중 표준코드 매핑 실패(X9999) 계정에 AI 1차 분류 제안을 붙인다."""
    unmapped = sorted(
        {
            str(row.get("account_name") or "").strip()
            for row in rows
            if str(row.get("account_name") or "").strip()
            and normalize_account_name(str(row.get("account_name") or ""), REFERENCE.aliases) == "other"
        }
    )
    result = call_ai_classification(unmapped, session)
    suggestions = result.get("suggestions") or {}
    for row in rows:
        suggestion = suggestions.get(str(row.get("account_name") or "").strip())
        if suggestion:
            row["ai_suggestion"] = suggestion
    return rows, result


def apply_ai_decisions(rows: list[dict], decisions: dict | None) -> tuple[list[dict], dict]:
    """추출 반영 시 AI 1차 분류 제안에 대한 계정별 승인/거절 결정(1차 승인)을 적용한다.

    decisions가 전달되면 명시적으로 approved된 계정의 제안만 유지하고,
    거절되거나 결정되지 않은 제안은 제거해 X9999 담당자 분류 대상으로 남긴다.
    decisions가 None이면(구버전 클라이언트) 기존처럼 제안 전체를 적용한다.
    """
    summary = {"approved": [], "rejected": [], "undecided": [], "per_account_review": decisions is not None}
    prepared = []
    for row in rows:
        row = dict(row)
        suggestion = row.get("ai_suggestion")
        if suggestion:
            account_name = str(row.get("account_name") or "").strip()
            if decisions is None:
                summary["approved"].append(account_name)
            else:
                decision = str(decisions.get(account_name) or "").strip().lower()
                if decision == "approved":
                    summary["approved"].append(account_name)
                elif decision == "rejected":
                    summary["rejected"].append(account_name)
                    row.pop("ai_suggestion", None)
                else:
                    summary["undecided"].append(account_name)
                    row.pop("ai_suggestion", None)
        prepared.append(row)
    return prepared, summary


def call_gemini_ocr(file_path: Path, mime_type: str, model: str) -> tuple[list[dict], list[str]]:
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        return [], ["GEMINI_API_KEY가 서버 환경변수에 설정되지 않았습니다."]
    file_bytes = file_path.read_bytes()
    if len(file_bytes) > GEMINI_INLINE_LIMIT_BYTES:
        return [], [f"파일이 {GEMINI_INLINE_LIMIT_BYTES // (1024 * 1024)}MB를 초과해 inline OCR 처리 대상에서 제외되었습니다."]

    prompt = """
한국 K-GAAP 재무제표 원본에서 계정명과 금액을 추출하세요.
반드시 JSON만 반환하세요. 설명 문장, 마크다운, 코드블록은 금지합니다.
형식:
{
  "rows": [
    {"account_name": "현금및현금성자산", "amount": 120000000}
  ],
  "issues": ["검토가 필요한 사항"]
}
규칙:
- 재무상태표, 손익계산서, 현금흐름표의 주요 계정과 금액을 추출합니다.
- 금액은 숫자만 반환하고 쉼표, 원, 천원, 백만원 단위 표기는 제거합니다.
- 괄호 금액은 음수로 반환합니다.
- 표가 흐리거나 계정/금액 판단이 불확실하면 issues에 적습니다.
- 계정명이 없거나 합계 제목만 있는 행은 제외합니다.
""".strip()
    input_type = "document" if mime_type == "application/pdf" else "image"
    payload = {
        "model": model,
        "store": False,
        "input": [
            {
                "type": input_type,
                "data": base64.b64encode(file_bytes).decode("ascii"),
                "mime_type": mime_type,
            },
            {"type": "text", "text": prompt},
        ],
        "response_format": {
            "type": "text",
            "mime_type": "application/json",
            "schema": {
                "type": "object",
                "properties": {
                    "rows": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "account_name": {"type": "string"},
                                "amount": {"type": "number"},
                            },
                            "required": ["account_name", "amount"],
                        },
                    },
                    "issues": {"type": "array", "items": {"type": "string"}},
                },
                "required": ["rows"],
            },
        },
    }
    endpoint = "https://generativelanguage.googleapis.com/v1beta/interactions"
    try:
        response = requests.post(endpoint, json=payload, headers={"x-goog-api-key": api_key}, timeout=45)
        response.raise_for_status()
        response_payload = response.json()
    except requests.HTTPError as exc:
        message = exc.response.text[:500]
        return [], [f"Gemini OCR 요청 실패: HTTP {exc.response.status_code}", message]
    except requests.Timeout:
        return [], ["Gemini OCR 요청 시간이 초과되었습니다."]
    except requests.RequestException as exc:
        return [], [f"Gemini OCR 네트워크 오류: {exc}"]

    text = gemini_response_text(response_payload)
    if not text:
        response_keys = ", ".join(sorted(response_payload.keys())) or "없음"
        return [], [f"Gemini OCR 응답에서 추출 텍스트를 찾지 못했습니다. 응답 키: {response_keys}"]
    try:
        parsed_payload = extract_json_object(text)
    except json.JSONDecodeError:
        return [], ["Gemini OCR 응답을 JSON으로 해석하지 못했습니다.", text[:500]]

    rows = parse_statement_rows(parsed_payload)
    issues = parsed_payload.get("issues") if isinstance(parsed_payload.get("issues"), list) else []
    issues = [str(issue) for issue in issues if str(issue).strip()]
    if not rows:
        issues.append("Gemini OCR은 실행됐지만 계정/금액 행을 찾지 못했습니다.")
    return rows, issues


def upload_file_path(upload: dict) -> Path:
    stored_path = UPLOAD_DIR / upload["stored_name"]
    if stored_path.exists():
        return stored_path
    file_bytes = upload.get("file_bytes")
    if isinstance(file_bytes, memoryview):
        file_bytes = file_bytes.tobytes()
    if isinstance(file_bytes, bytes):
        stored_path.write_bytes(file_bytes)
        return stored_path
    raise FileNotFoundError("업로드 원본 파일을 로컬 디스크 또는 DB에서 찾지 못했습니다.")


def extract_rows_from_upload(upload: dict) -> tuple[list[dict], list[str], str]:
    try:
        stored_path = upload_file_path(upload)
    except FileNotFoundError as exc:
        return [], [str(exc)], "missing_upload_file"
    suffix = Path(upload["original_name"]).suffix.lower()
    content_type = upload["content_type"].lower()
    config = ocr_config()
    if suffix == ".csv" or "csv" in content_type:
        text = stored_path.read_text(encoding="utf-8-sig")
        rows = parse_statement_rows({"csv_text": text})
        issues = [] if rows else ["CSV 파일은 읽었지만 계정 행을 찾지 못했습니다."]
        return rows, issues, "local_csv_parser"

    if suffix == ".xlsx" or "spreadsheetml" in content_type:
        try:
            rows, issues = parse_xlsx_statement_rows(stored_path)
        except (KeyError, ValueError, zipfile.BadZipFile, InvalidFileException) as exc:
            rows, issues = [], [f"Excel 파일 구조를 해석하지 못했습니다: {exc}"]
        return rows, issues, "local_xlsx_parser"

    if suffix == ".xls":
        return [], ["구형 .xls 파일은 아직 지원하지 않습니다. .xlsx 또는 CSV로 변환해 업로드하세요."], "unsupported_excel"

    ocr_mime = supported_ocr_mime(Path(upload["original_name"]), content_type)
    pdf_table_issues: list[str] = []
    if suffix == ".pdf" or ocr_mime == "application/pdf":
        rows, pdf_table_issues = parse_pdf_statement_rows(stored_path)
        if rows:
            return rows, pdf_table_issues, "local_pdf_table_parser"
        if pdf_table_issues:
            pdf_table_issues = [*pdf_table_issues, "PDF 표 직접 추출 실패 후 OCR 분석으로 전환합니다."]

    if config["api_key_ready"] and config["provider"] == "gemini" and ocr_mime:
        rows, issues = call_gemini_ocr(stored_path, ocr_mime, config["model"])
        issues = [*pdf_table_issues, f"OCR 제공자: {config['provider']} / 모델: {config['model']}", *issues]
        return rows, issues, "gemini_ocr"

    if ocr_mime:
        return (
            [],
            [
                *pdf_table_issues,
                f"OCR 제공자: {config['provider']} / 모델: {config['model']}",
                "서버 OCR 키가 설정되지 않아 실제 PDF/이미지 분석을 실행하지 못했습니다.",
                "관리자가 서버 환경변수 GEMINI_API_KEY를 설정한 뒤 다시 분석하세요.",
            ],
            "ocr_not_configured",
        )

    sample_rows = [
        {"account_name": "현금및현금성자산", "amount": 120000000},
        {"account_name": "매출채권", "amount": 88000000},
        {"account_name": "리스부채", "amount": 30000000},
        {"account_name": "개발비", "amount": 45000000},
    ]
    issues = [
        f"OCR 제공자: {config['provider']} / 모델: {config['model']}",
        "GEMINI_API_KEY가 설정되지 않아 실제 OCR 대신 샘플 추출을 사용했습니다." if not config["api_key_ready"] else "이 파일 형식은 현재 OCR 처리 대상이 아니어서 샘플 추출을 사용했습니다.",
        "워크플로우 확인을 위해 샘플 추출 행을 생성했습니다.",
        "변환 전에 담당자가 추출 결과를 교체하거나 확인해야 합니다.",
    ]
    return sample_rows, issues, "ocr_placeholder"



# ───────────────────────────────────────────────────────────────────────────
# §10 FastAPI 앱: 의존성 → 요청 모델 → 라우트 → 정적 서빙 → main
# ───────────────────────────────────────────────────────────────────────────

@contextlib.asynccontextmanager
async def lifespan(_app: FastAPI):
    """서버 시작 시 DB 스키마·기준정보 시드·계약 검증을 수행한다 (실패하면 기동 중단)."""
    init_db()
    yield


app = FastAPI(title="GTF K-GAAP → K-IFRS 변환", lifespan=lifespan)


@app.exception_handler(HTTPException)
async def flat_error_handler(_request: Request, exc: HTTPException):
    """에러 응답을 {"detail": ...}가 아니라 기존 프런트엔드가 기대하는 평평한 JSON으로 유지한다."""
    body = exc.detail if isinstance(exc.detail, dict) else {"error": str(exc.detail)}
    return JSONResponse(status_code=exc.status_code, content=body)




# --- 인증 의존성 (Depends) ---

def get_current_user(request: Request, session: Session) -> AppUser | None:
    """세션 쿠키 토큰으로 로그인 사용자를 조회한다. 없으면 None."""
    token = request.cookies.get(SESSION_COOKIE, "")
    if not token:
        return None
    return session.scalar(
        select(AppUser)
        .join(UserSession, UserSession.user_id == AppUser.id)
        .where(UserSession.token_hash == session_token_hash(token), UserSession.expires_at > utc_now())
    )


def current_user(request: Request, session: Session = Depends(get_db)) -> AppUser | None:
    return get_current_user(request, session)


def require_user(user: AppUser | None = Depends(current_user)) -> AppUser:
    """로그인 필수 라우트의 의존성. 미로그인이면 401."""
    if not user:
        raise HTTPException(401, {"error": "로그인이 필요합니다.", "login_required": True})
    return user


def create_login_session(session: Session, user_id: str) -> str:
    """세션 토큰을 만들어 해시만 저장하고 원본 토큰(쿠키에 넣을 값)을 돌려준다."""
    token = secrets.token_urlsafe(32)
    expires_at = datetime.fromtimestamp(
        datetime.now(timezone.utc).timestamp() + SESSION_MAX_AGE_SECONDS, tz=timezone.utc
    ).isoformat()
    session.add(
        UserSession(
            id=str(uuid.uuid4()),
            user_id=user_id,
            token_hash=session_token_hash(token),
            created_at=utc_now(),
            expires_at=expires_at,
        )
    )
    session.commit()
    return token


def require_write_user(user: AppUser = Depends(require_user)) -> AppUser:
    """쓰기 라우트의 의존성. 읽기 전용 데모 계정이면 403."""
    if user.is_read_only:
        raise HTTPException(403, {"error": "읽기 전용 데모 계정에서는 데이터를 생성, 수정, 삭제할 수 없습니다.", "read_only": True})
    return user


def set_session_cookie(response: Response, token: str) -> None:
    response.set_cookie(SESSION_COOKIE, token, max_age=SESSION_MAX_AGE_SECONDS, path="/", httponly=True, samesite="lax")


def get_project_or_404(session: Session, project_id: str, owner_user_id: str | None = None) -> Project:
    """프로젝트를 조회한다. 없으면(또는 소유자가 다르면) 404."""
    stmt = select(Project).where(Project.id == project_id)
    if owner_user_id is not None:
        stmt = stmt.where(Project.owner_user_id == owner_user_id)
    project = session.scalar(stmt)
    if not project:
        raise HTTPException(404, {"error": "Project not found"})
    return project


# --- 요청 본문 모델 (pydantic) ---

class LoginRequest(BaseModel):
    email: str = ""
    password: str = ""


class ProjectCreateRequest(BaseModel):
    company_name: str = ""
    source_standard: str = ""
    target_standard: str = ""
    period: str = ""


class StatementsAddRequest(BaseModel):
    rows: list[dict] | None = None
    csv_text: str = ""
    source: str = "manual"


class AcceptExtractionRequest(BaseModel):
    ai_decisions: dict[str, str] | None = None


class ConvertRequest(BaseModel):
    responses: dict[str, dict] = {}


class ClassifyStatementRequest(BaseModel):
    account_key: str = ""


class ReviewRequest(BaseModel):
    decision: str = ""
    reviewer_name: str = ""
    memo: str = ""


# --- 헬스체크·설정 조회 ---

@app.get("/healthz")
def healthz():
    return {
        "status": "ok",
        "service": "gtf-accounting-conversion",
        "time": utc_now(),
        "database": database_ready(),
        "database_config": database_config(),
        "ocr": ocr_config(),
        "ai": ai_config(),
        "dart": dart_config(),
    }


@app.get("/api/ocr-config")
def get_ocr_config(user: AppUser = Depends(require_user)):
    return ocr_config()


@app.get("/api/ai-config")
def get_ai_config(user: AppUser = Depends(require_user)):
    return ai_config()


@app.get("/api/dart-config")
def get_dart_config(user: AppUser = Depends(require_user)):
    return dart_config()


# --- 인증 ---

@app.get("/api/auth/session")
def auth_session(user: AppUser | None = Depends(current_user)):
    return {
        "authenticated": bool(user),
        "user": user_public_dict(row_to_dict(user)) if user else None,
        "admin_configured": admin_config()["configured"],
    }


@app.post("/api/auth/login")
def login(payload: LoginRequest, response: Response, session: Session = Depends(get_db)):
    email = normalize_email(payload.email)
    if not email or not payload.password:
        raise HTTPException(400, {"error": "이메일과 비밀번호를 입력하세요."})
    ensure_admin_user(session)
    user = session.scalar(select(AppUser).where(AppUser.email == email))
    if not user and not admin_config()["configured"]:
        raise HTTPException(503, {"error": "관리자 계정이 설정되지 않았습니다. ADMIN_EMAIL과 ADMIN_PASSWORD를 서버 환경변수에 설정하세요."})
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(401, {"error": "이메일 또는 비밀번호가 올바르지 않습니다."})
    token = create_login_session(session, user.id)
    set_session_cookie(response, token)
    return {"authenticated": True, "user": user_public_dict(row_to_dict(user))}


@app.post("/api/auth/demo")
def demo_login(response: Response, session: Session = Depends(get_db)):
    ensure_admin_user(session)
    user = ensure_demo_user(session)
    if not user:
        raise HTTPException(403, {"error": "데모 로그인이 비활성화되어 있습니다."})
    token = create_login_session(session, user["id"])
    set_session_cookie(response, token)
    return {"authenticated": True, "user": user_public_dict(user), "demo": True}


@app.post("/api/auth/logout")
def logout(request: Request, response: Response, session: Session = Depends(get_db)):
    token = request.cookies.get(SESSION_COOKIE, "")
    if token:
        session.execute(delete(UserSession).where(UserSession.token_hash == session_token_hash(token)))
        session.commit()
    response.delete_cookie(SESSION_COOKIE, path="/")
    return {"authenticated": False, "user": None}


# --- 기준정보·기준서 검색 ---

REFERENCE_TABLE_LABELS = [
    (StandardAccount, "내부 표준계정코드 DB"),
    (KgaapAccount, "K-GAAP 계정명 DB"),
    (ChecklistItem, "판단 체크리스트 DB"),
    (StandardsParagraph, "K-GAAP/K-IFRS 기준서 문단 검색 DB"),
    (FinancialStatementTemplate, "재무제표 양식 DB"),
]


@app.get("/api/reference-data")
def reference_data(user: AppUser = Depends(require_user), session: Session = Depends(get_db)):
    summary = [
        {
            "table": model.__tablename__,
            "label": label,
            "count": session.scalar(select(func.count()).select_from(model)),
        }
        for model, label in REFERENCE_TABLE_LABELS
    ]
    accounts = [
        dict(row._mapping)
        for row in session.execute(
            select(
                StandardAccount.account_key,
                StandardAccount.standard_code,
                StandardAccount.internal_label,
                StandardAccount.ifrs_account,
                StandardAccount.mapping_type,
            ).order_by(StandardAccount.standard_code)
        ).all()
    ]
    templates = [
        dict(row._mapping)
        for row in session.execute(
            select(
                FinancialStatementTemplate.statement_type,
                FinancialStatementTemplate.section,
                FinancialStatementTemplate.line_item,
                FinancialStatementTemplate.account_key,
                FinancialStatementTemplate.display_order,
            )
            .where(FinancialStatementTemplate.standard_set == "IFRS", FinancialStatementTemplate.active.is_(True))
            .order_by(FinancialStatementTemplate.statement_type, FinancialStatementTemplate.display_order)
        ).all()
    ]
    return {"summary": summary, "accounts": accounts, "templates": templates}


@app.get("/api/standards/search")
def standards_search(
    q: str = "",
    account_key: str = "",
    standard_set: str = "",
    user: AppUser = Depends(require_user),
    session: Session = Depends(get_db),
):
    query, account_key, standard_set = q.strip(), account_key.strip(), standard_set.strip()
    if standard_set and standard_set not in {"K-GAAP", "K-IFRS"}:
        raise HTTPException(400, {"error": "standard_set은 K-GAAP 또는 K-IFRS여야 합니다."})
    if query:
        paragraphs = semantic_search_paragraphs(
            session, query, account_key=account_key or None, standard_set=standard_set or None, k=8
        )
    else:
        paragraphs = find_standards_paragraphs(
            session, account_key=account_key or None, query=None, standard_set=standard_set or None
        )
    return {
        "count": len(paragraphs),
        "retrieval": paragraphs[0].get("retrieval", "none") if paragraphs else "none",
        "standard_sets": ["K-GAAP", "K-IFRS"],
        "paragraphs": paragraphs,
        "note": "기준서 문단 요약 기준정보입니다. 최종 판단 시 기준서 원문을 확인하세요.",
    }


# --- 프로젝트 ---

@app.get("/api/projects")
def list_projects(user: AppUser = Depends(require_user), session: Session = Depends(get_db)):
    projects = session.scalars(
        select(Project).where(Project.owner_user_id == user.id).order_by(Project.created_at.desc())
    )
    return [row_to_dict(project) for project in projects]


@app.post("/api/projects", status_code=201)
def create_project(
    payload: ProjectCreateRequest,
    user: AppUser = Depends(require_write_user),
    session: Session = Depends(get_db),
):
    now = utc_now()
    project = Project(
        id=str(uuid.uuid4()),
        owner_user_id=user.id,
        is_test=False,
        company_name=payload.company_name or "Untitled company",
        source_standard=payload.source_standard or "K-GAAP",
        target_standard=payload.target_standard or "IFRS",
        period=payload.period or "2026",
        status="created",
        created_at=now,
        updated_at=now,
    )
    session.add(project)
    session.flush()  # audit_logs의 project_id 외래키가 유효하도록 먼저 반영
    payload_dict = row_to_dict(project)
    log_event(session, project.id, "project.created", payload_dict)
    session.commit()
    return payload_dict


@app.get("/api/projects/{project_id}")
def get_project(project_id: str, user: AppUser = Depends(require_user), session: Session = Depends(get_db)):
    project = get_project_or_404(session, project_id, owner_user_id=user.id)
    statements = load_project_statements(session, project_id)
    uploads = [
        upload_public_dict(upload)
        for upload in session.scalars(
            select(Upload).where(Upload.project_id == project_id).order_by(Upload.created_at.desc())
        )
    ]
    extractions = [
        dict(
            row_to_dict(extraction),
            rows=parse_json_field(extraction.rows_json, []),
            issues=parse_json_field(extraction.issues_json, []),
        )
        for extraction in session.scalars(
            select(Extraction).where(Extraction.project_id == project_id).order_by(Extraction.created_at.desc())
        )
    ]
    conversion = session.scalar(
        select(Conversion).where(Conversion.project_id == project_id).order_by(Conversion.created_at.desc()).limit(1)
    )
    review = session.scalar(
        select(Review).where(Review.project_id == project_id).order_by(Review.created_at.desc()).limit(1)
    )
    return {
        "project": row_to_dict(project),
        "statements": statements,
        "uploads": uploads,
        "extractions": extractions,
        "conversion": parse_json_field(conversion.output_json, None) if conversion else None,
        "review": row_to_dict(review) if review else None,
    }


@app.delete("/api/projects/{project_id}")
def delete_project(project_id: str, user: AppUser = Depends(require_write_user), session: Session = Depends(get_db)):
    get_project_or_404(session, project_id, owner_user_id=user.id)
    stored_names = list(session.scalars(select(Upload.stored_name).where(Upload.project_id == project_id)))
    # 외래키 순서: 자식 테이블부터 지우고 마지막에 프로젝트를 지운다.
    for model in (Extraction, Upload, Statement, Conversion, Review, AuditLog):
        session.execute(delete(model).where(model.project_id == project_id))
    session.execute(delete(Project).where(Project.id == project_id, Project.owner_user_id == user.id))
    session.commit()
    for stored_name in stored_names:
        if stored_name:
            try:
                (UPLOAD_DIR / stored_name).unlink(missing_ok=True)
            except OSError:
                pass
    return {"deleted": True, "project_id": project_id}


# --- 업로드·추출 ---

@app.get("/api/projects/{project_id}/uploads")
def list_uploads(project_id: str, user: AppUser = Depends(require_user), session: Session = Depends(get_db)):
    uploads = session.scalars(
        select(Upload).where(Upload.project_id == project_id).order_by(Upload.created_at.desc())
    )
    return [upload_public_dict(upload) for upload in uploads]


@app.post("/api/projects/{project_id}/uploads", status_code=201)
def upload_file(
    project_id: str,
    file: UploadFile = File(...),
    user: AppUser = Depends(require_write_user),
    session: Session = Depends(get_db),
):
    get_project_or_404(session, project_id)

    original_name = file.filename or "upload.bin"
    content = file.file.read()
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(original_name).name).strip("._")
    stored_name = f"{project_id}_{uuid.uuid4()}_{safe_name or 'upload.bin'}"
    (UPLOAD_DIR / stored_name).write_bytes(content)

    upload = Upload(
        id=str(uuid.uuid4()),
        project_id=project_id,
        original_name=original_name,
        stored_name=stored_name,
        content_type=file.content_type or "application/octet-stream",
        size_bytes=len(content),
        file_bytes=content,
        extraction_status="pending_ocr",
        created_at=utc_now(),
    )
    session.add(upload)
    session.execute(update(Project).where(Project.id == project_id).values(status="source_uploaded", updated_at=utc_now()))
    log_event(
        session,
        project_id,
        "source.uploaded",
        {
            "upload_id": upload.id,
            "original_name": original_name,
            "content_type": upload.content_type,
            "size_bytes": len(content),
            "next_step": "Gemini OCR extraction",
        },
    )
    public = upload_public_dict(upload)
    session.commit()
    return public


@app.delete("/api/projects/{project_id}/uploads/{upload_id}")
def delete_upload(
    project_id: str,
    upload_id: str,
    user: AppUser = Depends(require_write_user),
    session: Session = Depends(get_db),
):
    upload = session.scalar(select(Upload).where(Upload.id == upload_id, Upload.project_id == project_id))
    if not upload:
        raise HTTPException(404, {"error": "Upload not found"})
    stored_name = upload.stored_name
    original_name = upload.original_name

    session.execute(delete(Extraction).where(Extraction.upload_id == upload_id, Extraction.project_id == project_id))
    session.delete(upload)
    session.flush()
    remaining = session.scalar(select(func.count()).select_from(Upload).where(Upload.project_id == project_id))
    next_status = "created" if int(remaining or 0) == 0 else "source_uploaded"
    session.execute(update(Project).where(Project.id == project_id).values(status=next_status, updated_at=utc_now()))
    log_event(
        session,
        project_id,
        "source.deleted",
        {"upload_id": upload_id, "original_name": original_name, "remaining_uploads": remaining},
    )
    session.commit()
    if stored_name:
        try:
            (UPLOAD_DIR / stored_name).unlink(missing_ok=True)
        except OSError:
            pass
    return {"deleted": True, "upload_id": upload_id, "project_status": next_status}


@app.get("/api/projects/{project_id}/extractions")
def list_extractions(project_id: str, user: AppUser = Depends(require_user), session: Session = Depends(get_db)):
    extractions = session.scalars(
        select(Extraction).where(Extraction.project_id == project_id).order_by(Extraction.created_at.desc())
    )
    return [
        dict(
            row_to_dict(extraction),
            rows=parse_json_field(extraction.rows_json, []),
            issues=parse_json_field(extraction.issues_json, []),
        )
        for extraction in extractions
    ]


@app.post("/api/projects/{project_id}/uploads/{upload_id}/extract", status_code=201)
def extract_upload(
    project_id: str,
    upload_id: str,
    user: AppUser = Depends(require_write_user),
    session: Session = Depends(get_db),
):
    upload = session.scalar(select(Upload).where(Upload.id == upload_id, Upload.project_id == project_id))
    if not upload:
        raise HTTPException(404, {"error": "Upload not found"})

    config = ocr_config()
    rows, issues, provider = extract_rows_from_upload(row_to_dict(upload))
    rows, ai_classification = attach_ai_classification(rows, session)
    if ai_classification.get("status") != "skipped" and ai_classification.get("note"):
        issues = [*issues, ai_classification["note"]]
    status = "needs_review" if rows else "failed"
    extraction = Extraction(
        id=str(uuid.uuid4()),
        project_id=project_id,
        upload_id=upload_id,
        provider=provider,
        status=status,
        rows_json=json.dumps(rows, ensure_ascii=False),
        issues_json=json.dumps(issues, ensure_ascii=False),
        created_at=utc_now(),
    )
    session.add(extraction)
    upload.extraction_status = status
    session.execute(update(Project).where(Project.id == project_id).values(status="extracted", updated_at=utc_now()))
    log_event(
        session,
        project_id,
        "source.extracted",
        {
            "upload_id": upload_id,
            "extraction_id": extraction.id,
            "provider": provider,
            "ocr_config": config,
            "row_count": len(rows),
            "issues": issues,
            "ai_classification": ai_classification_audit(ai_classification),
        },
    )
    result = dict(row_to_dict(extraction), rows=rows, issues=issues)
    session.commit()
    return result


# --- DART 연동 ---

@app.post("/api/projects/{project_id}/dart/import")
def dart_import(
    project_id: str,
    payload: dict = Body(default={}),
    user: AppUser = Depends(require_write_user),
    session: Session = Depends(get_db),
):
    get_project_or_404(session, project_id)

    rows, issues, metadata = fetch_dart_statement_rows(payload, REFERENCE.aliases)
    rows, ai_classification = attach_ai_classification(rows, session)
    if ai_classification.get("status") != "skipped" and ai_classification.get("note"):
        issues = [*issues, ai_classification["note"]]
    raw_rows = metadata.pop("raw_rows", [])
    raw_payload = {"metadata": metadata, "raw_rows": raw_rows, "filtered_rows": rows, "issues": issues}
    raw_bytes = json.dumps(raw_payload, ensure_ascii=False).encode("utf-8")
    status = "needs_review" if rows else "failed"
    now = utc_now()

    upload = Upload(
        id=str(uuid.uuid4()),
        project_id=project_id,
        original_name=f"DART_API_{metadata.get('corp_code', 'unknown')}_{metadata.get('bsns_year', payload.get('bsns_year', 'unknown'))}.json",
        stored_name="",
        content_type="application/json",
        size_bytes=len(raw_bytes),
        file_bytes=raw_bytes,
        extraction_status=status,
        created_at=now,
    )
    session.add(upload)
    session.flush()
    extraction = Extraction(
        id=str(uuid.uuid4()),
        project_id=project_id,
        upload_id=upload.id,
        provider="dart_api",
        status=status,
        rows_json=json.dumps(rows, ensure_ascii=False),
        issues_json=json.dumps([*issues, json.dumps(metadata, ensure_ascii=False)], ensure_ascii=False),
        created_at=now,
    )
    session.add(extraction)
    next_status = "extracted" if rows else "source_import_failed"
    session.execute(update(Project).where(Project.id == project_id).values(status=next_status, updated_at=utc_now()))
    log_event(
        session,
        project_id,
        "dart.imported",
        {
            "upload_id": upload.id,
            "extraction_id": extraction.id,
            "row_count": len(rows),
            "raw_row_count": len(raw_rows),
            "issues": issues,
            "metadata": metadata,
            "ai_classification": ai_classification_audit(ai_classification),
        },
    )
    body = {
        **row_to_dict(extraction),
        "rows": rows,
        "issues": issues,
        "metadata": metadata,
        "upload": upload_public_dict(upload),
    }
    session.commit()
    return JSONResponse(status_code=201 if rows else 400, content=body)


@app.post("/api/projects/{project_id}/dart/reports")
def dart_reports(
    project_id: str,
    payload: dict = Body(default={}),
    user: AppUser = Depends(require_write_user),
    session: Session = Depends(get_db),
):
    get_project_or_404(session, project_id)
    reports, issues, metadata = fetch_dart_available_reports(payload)
    return {"reports": reports, "issues": issues, "metadata": metadata}


# --- 추출 반영(1차 승인)·수동 입력·검증 ---

@app.post("/api/projects/{project_id}/extractions/{extraction_id}/accept")
def accept_extraction(
    project_id: str,
    extraction_id: str,
    payload: AcceptExtractionRequest | None = None,
    user: AppUser = Depends(require_write_user),
    session: Session = Depends(get_db),
):
    ai_decisions = payload.ai_decisions if payload else None
    project = session.get(Project, project_id)
    extraction = session.scalar(
        select(Extraction).where(Extraction.id == extraction_id, Extraction.project_id == project_id)
    )
    if not project or not extraction:
        raise HTTPException(404, {"error": "Extraction not found"})

    rows = parse_json_field(extraction.rows_json, [])
    rows, decision_summary = apply_ai_decisions(rows, ai_decisions)
    records = [build_statement_record(project.period, row, REFERENCE) for row in rows]
    for record in records:
        session.add(
            Statement(
                id=record["id"],
                project_id=project_id,
                account_name=record["account_name"],
                normalized_account=record["normalized_account"],
                standard_code=record["standard_code"],
                amount=record["amount"],
                period=record["period"],
                mapping_type=record["mapping_type"],
                rule_summary=record["rule_summary"],
                checklist_json=json.dumps(record["checklist"], ensure_ascii=False),
                created_at=utc_now(),
            )
        )
    extraction.status = "accepted"
    project.status = "mapped"
    project.updated_at = utc_now()
    ai_confirmed = [
        {
            "account_name": record["account_name"],
            "suggested_account": record["normalized_account"],
            "confidence": (record.get("ai_suggestion") or {}).get("confidence"),
            "rationale": (record.get("ai_suggestion") or {}).get("rationale"),
        }
        for record in records
        if record.get("mapping_source") == "ai_suggested_human_accepted"
    ]
    log_event(
        session,
        project_id,
        "extraction.accepted",
        {
            "extraction_id": extraction_id,
            "statement_count": len(records),
            "ai_classified_count": len(ai_confirmed),
            "ai_classified_accounts": ai_confirmed,
            "ai_decision_summary": decision_summary,
            "ai_classification_note": (
                "AI 1차 분류 제안을 담당자가 계정별로 승인/거절하며 확정했습니다."
                if decision_summary["per_account_review"] and (ai_confirmed or decision_summary["rejected"])
                else "AI 1차 분류 제안을 담당자가 반영하며 확정했습니다." if ai_confirmed else None
            ),
        },
        actor=user.email or "system",
    )
    session.commit()
    return {"statements": records, "extraction_id": extraction_id, "ai_decision_summary": decision_summary}


@app.post("/api/projects/{project_id}/statements", status_code=201)
def add_statements(
    project_id: str,
    payload: StatementsAddRequest,
    user: AppUser = Depends(require_write_user),
    session: Session = Depends(get_db),
):
    raw_rows = parse_statement_rows(payload.model_dump())
    # 수동 입력도 파일/DART 경로와 동일하게 추출(extraction)로 만들어, 미분류 계정에
    # AI 1차 분류 제안을 붙이고 담당자가 추출 미리보기에서 계정별로 승인하도록 통일한다.
    raw_rows, ai_classification = attach_ai_classification(raw_rows, session)
    issues = []
    if ai_classification.get("status") != "skipped" and ai_classification.get("note"):
        issues.append(ai_classification["note"])
    status = "needs_review" if raw_rows else "failed"
    now = utc_now()

    get_project_or_404(session, project_id)
    upload = Upload(
        id=str(uuid.uuid4()),
        project_id=project_id,
        original_name="수동입력.csv",
        stored_name="",
        content_type="text/csv",
        size_bytes=0,
        extraction_status=status,
        created_at=now,
    )
    session.add(upload)
    session.flush()
    extraction = Extraction(
        id=str(uuid.uuid4()),
        project_id=project_id,
        upload_id=upload.id,
        provider="manual_input",
        status=status,
        rows_json=json.dumps(raw_rows, ensure_ascii=False),
        issues_json=json.dumps(issues, ensure_ascii=False),
        created_at=now,
    )
    session.add(extraction)
    session.execute(update(Project).where(Project.id == project_id).values(status="extracted", updated_at=now))
    log_event(
        session,
        project_id,
        "source.manual_entered",
        {
            "extraction_id": extraction.id,
            "row_count": len(raw_rows),
            "source": payload.source,
            "ai_classification": ai_classification_audit(ai_classification),
        },
    )
    extraction_id = extraction.id
    session.commit()
    return {
        "extraction_id": extraction_id,
        "rows": raw_rows,
        "issues": issues,
        "ai_classification_status": ai_classification.get("status"),
    }


@app.patch("/api/projects/{project_id}/statements/{statement_id}/classify")
def classify_statement(
    project_id: str,
    statement_id: str,
    payload: ClassifyStatementRequest,
    user: AppUser = Depends(require_write_user),
    session: Session = Depends(get_db),
):
    """반영된 계정 행을 담당자가 표준계정으로 재분류한다.

    검토 요약의 '미분류(X9999)' 오류 항목이 유도하는 행동. 재분류 전/후가 감사 로그에 남아
    'AI 제안 → 사람 확정'과 같은 원칙(분류 확정 권한은 사람, 과정은 기록)을 따른다.
    """
    get_project_or_404(session, project_id)
    statement = session.get(Statement, statement_id)
    if not statement or statement.project_id != project_id:
        raise HTTPException(404, {"error": "계정 행을 찾지 못했습니다."})
    account_key = payload.account_key.strip()
    account = REFERENCE.accounts.get(account_key)
    if not account or account_key == "other":
        raise HTTPException(400, {"error": "유효한 표준계정 키가 아닙니다. 분류 가능한 계정 목록에서 선택하세요."})
    before = {"standard_code": statement.standard_code, "normalized_account": statement.normalized_account}
    checklist = REFERENCE.checklists.get(account_key, []) if account["type"] == "judgment" else []
    statement.normalized_account = account["label"]
    statement.standard_code = account["code"]
    statement.mapping_type = account["type"]
    statement.rule_summary = f"[담당자 재분류] {account['rule']}"
    statement.checklist_json = json.dumps(checklist, ensure_ascii=False)
    log_event(
        session,
        project_id,
        "statement.reclassified",
        {
            "statement_id": statement_id,
            "account_name": statement.account_name,
            "before": before,
            "after": {"standard_code": account["code"], "normalized_account": account["label"]},
        },
        actor=user.email,
    )
    session.commit()
    return row_to_dict(statement)


@app.post("/api/projects/{project_id}/validate")
def validate_project(project_id: str, user: AppUser = Depends(require_write_user), session: Session = Depends(get_db)):
    project = get_project_or_404(session, project_id)
    statements = [row_to_dict(s) for s in session.scalars(select(Statement).where(Statement.project_id == project_id))]
    result = validate_statement_records(row_to_dict(project), statements)
    log_event(session, project_id, "validation.completed", result)
    session.commit()
    return result


# --- 변환·검토 ---

@app.post("/api/projects/{project_id}/policy-comparison")
def policy_comparison(
    project_id: str,
    payload: ConvertRequest,
    user: AppUser = Depends(require_user),
    session: Session = Depends(get_db),
):
    """선택가능 회계정책(원가/재평가, 원가/공정가치, 자산차감/이연수익)의 영향 비교.

    결정론 계산기를 선택지별 입력으로 재실행하는 조회성 산출 — 저장·확정 없음(읽기 전용도 허용).
    """
    project = get_project_or_404(session, project_id)
    statements = load_project_statements(session, project_id)
    return compare_policy_scenarios(row_to_dict(project), statements, payload.responses or {}, REFERENCE)


@app.post("/api/projects/{project_id}/convert")
def convert_project(
    project_id: str,
    payload: ConvertRequest,
    user: AppUser = Depends(require_write_user),
    session: Session = Depends(get_db),
):
    responses = payload.responses or {}
    project_row = get_project_or_404(session, project_id)
    project = row_to_dict(project_row)
    statement_rows = load_project_statements(session, project_id)
    output = generate_conversion(project, statement_rows, responses, REFERENCE)
    # RAG: 판단 필요 항목마다 계정명·근거를 질의로 관련 기준서 문단을 시맨틱 검색해
    # AI 판단 보조의 근거(context)로 주입한다. 검색 결과는 조정 금액이 아니라 근거 설명에만 쓰인다.
    retrieved_context = []
    for jitem in output["judgment_items"]:
        query = f"{jitem.get('account', '')} {jitem.get('basis', '')}".strip()
        paras = semantic_search_paragraphs(session, query, k=3) if query else []
        retrieved_context.append(
            {
                "account": jitem.get("account"),
                "paragraphs": [
                    {
                        "standard_set": p.get("standard_set"),
                        "reference_code": p.get("reference_code"),
                        "title": p.get("title"),
                        "content": p.get("content"),
                        "retrieval": p.get("retrieval"),
                        "similarity": p.get("similarity"),
                    }
                    for p in paras
                ],
            }
        )
    output["ai_assistance"] = call_ai_judgment(project, output["entries"], output["judgment_items"], retrieved_context)
    output["retrieved_context"] = retrieved_context

    session.add(
        Conversion(
            id=str(uuid.uuid4()),
            project_id=project_id,
            output_json=json.dumps(output, ensure_ascii=False),
            created_at=utc_now(),
        )
    )
    session.execute(update(Project).where(Project.id == project_id).values(status="draft_generated", updated_at=utc_now()))
    log_event(
        session,
        project_id,
        "conversion.generated",
        {
            "responses": responses,
            "entry_count": len(output["entries"]),
            "template": output["statement_template"],
            "ai_status": output["ai_assistance"].get("status"),
        },
    )
    session.commit()
    return output


@app.get("/api/projects/{project_id}/review-summary")
def review_summary(project_id: str, user: AppUser = Depends(require_user), session: Session = Depends(get_db)):
    project = get_project_or_404(session, project_id)
    statements = sort_statements_by_code([
        row_to_dict(statement)
        for statement in session.scalars(select(Statement).where(Statement.project_id == project_id))
    ])
    conversion_row = session.scalar(
        select(Conversion).where(Conversion.project_id == project_id).order_by(Conversion.created_at.desc()).limit(1)
    )
    validation = validate_statement_records(row_to_dict(project), statements) if statements else None
    conversion = parse_json_field(conversion_row.output_json, {}) if conversion_row else None
    return build_review_summary(statements, conversion, validation)


@app.post("/api/projects/{project_id}/review", status_code=201)
def record_review(
    project_id: str,
    payload: ReviewRequest,
    user: AppUser = Depends(require_write_user),
    session: Session = Depends(get_db),
):
    if payload.decision not in {"approved", "changes_requested"}:
        raise HTTPException(400, {"error": "Decision must be approved or changes_requested."})

    reviewer_name = payload.reviewer_name.strip() or "Unassigned reviewer"
    memo = payload.memo.strip()
    get_project_or_404(session, project_id)
    conversion = session.scalar(
        select(Conversion).where(Conversion.project_id == project_id).order_by(Conversion.created_at.desc()).limit(1)
    )
    if not conversion:
        raise HTTPException(400, {"error": "Generate a conversion draft before review."})
    if payload.decision == "approved":
        # 2차 승인 게이트: 오류 수준(미분류 잔존)은 승인을 차단, 경고는 검토자 판단에 맡긴다.
        unclassified = session.scalar(
            select(func.count())
            .select_from(Statement)
            .where(Statement.project_id == project_id, Statement.standard_code == "X9999")
        )
        if unclassified:
            raise HTTPException(
                409,
                {
                    "error": f"미분류 계정 {unclassified}건이 남아 있어 승인할 수 없습니다. 담당자 분류 또는 AI 제안 승인(1차 승인) 후 다시 시도하세요.",
                    "unclassified_count": unclassified,
                },
            )
    review = Review(
        id=str(uuid.uuid4()),
        project_id=project_id,
        reviewer_name=reviewer_name,
        decision=payload.decision,
        memo=memo,
        created_at=utc_now(),
    )
    session.add(review)
    session.execute(update(Project).where(Project.id == project_id).values(status=payload.decision, updated_at=utc_now()))
    log_event(
        session,
        project_id,
        "review.recorded",
        {"review_id": review.id, "decision": payload.decision, "memo": memo, "conversion_id": conversion.id},
        actor=reviewer_name,
    )
    body = row_to_dict(review)
    session.commit()
    return body


@app.get("/api/projects/{project_id}/audit")
def list_audit(project_id: str, user: AppUser = Depends(require_user), session: Session = Depends(get_db)):
    logs = session.scalars(
        select(AuditLog).where(AuditLog.project_id == project_id).order_by(AuditLog.created_at.desc())
    )
    return [dict(row_to_dict(log), detail=parse_json_field(log.detail_json, {})) for log in logs]


# --- 내보내기 ---

@app.get("/api/projects/{project_id}/exports/{export_name}")
def export_project(
    project_id: str,
    export_name: str,
    user: AppUser = Depends(require_user),
    session: Session = Depends(get_db),
):
    project = session.get(Project, project_id)
    if not project:
        raise HTTPException(404, {"error": "Project not found"})
    conversion = session.scalar(
        select(Conversion).where(Conversion.project_id == project_id).order_by(Conversion.created_at.desc()).limit(1)
    )
    if not conversion:
        raise HTTPException(400, {"error": "Generate a conversion draft before export."})

    output = parse_json_field(conversion.output_json, {})
    if export_name == "adjustments.csv":
        return Response(
            content=conversion_adjustments_csv(output).encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="gtf_adjustments.csv"'},
        )
    if export_name == "basis-report.txt":
        return Response(
            content=conversion_basis_report(output),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="gtf_basis_report.txt"'},
        )
    if export_name == "review-workbook.xlsx":
        statements = load_project_statements(session, project_id)
        latest = session.execute(
            select(Extraction.rows_json, Upload.file_bytes)
            .join(Upload, Upload.id == Extraction.upload_id, isouter=True)
            .where(Extraction.project_id == project_id)
            .order_by(Extraction.created_at.desc())
            .limit(1)
        ).first()
        audit_rows = [
            dict(row_to_dict(log), detail=parse_json_field(log.detail_json, {}))
            for log in session.scalars(
                select(AuditLog).where(AuditLog.project_id == project_id).order_by(AuditLog.created_at)
            )
        ]
        extraction_rows = dart_raw_rows_from_upload(dict(latest._mapping) if latest else None)
        if not extraction_rows and latest:
            extraction_rows = parse_json_field(latest.rows_json, [])
        workbook = review_workbook_bytes(row_to_dict(project), extraction_rows, statements, output, audit_rows)
        return Response(
            content=workbook,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="gtf_review_workbook.xlsx"'},
        )
    raise HTTPException(404, {"error": "Unknown export type"})


# --- 정적 파일(React 빌드)·SPA 진입점 — API 라우트보다 뒤에 두어야 catch-all이 API를 가리지 않는다 ---

@app.get("/styles.css")
def styles():
    return PlainTextResponse(STYLES_CSS, media_type="text/css")


@app.get("/app.js")
def script():
    return PlainTextResponse(APP_JS, media_type="application/javascript")


@app.get("/{full_path:path}")
def serve_frontend(full_path: str):
    """React 빌드 정적 파일을 서빙하고, 없으면 SPA 진입점(index.html)을 돌려준다."""
    static = figma_static_file("/" + full_path)
    if static is not None:
        cache = "no-cache" if static.name == "index.html" else "public, max-age=31536000, immutable"
        return FileResponse(static, headers={"Cache-Control": cache})
    if full_path.startswith("api/"):
        raise HTTPException(404, {"error": "Not found"})
    index = figma_static_file("/")
    return FileResponse(index, headers={"Cache-Control": "no-cache"}) if index else HTMLResponse(INDEX_HTML)


def main() -> None:
    port = int(os.environ.get("PORT", "4173"))
    host = os.environ.get("HOST", "127.0.0.1")
    uvicorn.run(app, host=host, port=port)


if __name__ == "__main__":
    main()
